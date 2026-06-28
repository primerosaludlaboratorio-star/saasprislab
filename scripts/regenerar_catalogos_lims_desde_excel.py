from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def _clean(value):
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r", "").replace("_x000D_", "")
    return text


def _rows_from_sheet(workbook_path: Path, sheet_name: str):
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        yield [_clean(cell) for cell in row]


def _trim_row(row):
    while row and row[-1] == "":
        row.pop()
    return row


def _write_csv_from_sheet(workbook_path: Path, sheet_name: str, output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        for row in _rows_from_sheet(workbook_path, sheet_name):
            writer.writerow(_trim_row(row))


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Regenera los CSV de datos_lims desde los Excel originales de PRISLAB."
        )
    )
    parser.add_argument(
        "--base-dir",
        default=r"C:\Users\jonil\Desktop\PRISLAB_SaaS-master\PRISLAB_SaaS-master",
        help="Ruta raíz del proyecto PRISLAB.",
    )
    parser.add_argument(
        "--downloads-dir",
        default=r"C:\Users\jonil\Downloads",
        help="Ruta donde están los Excel originales.",
    )
    parser.add_argument(
        "--tarifa",
        default="Tarifa_Detalle_20260602_010845.xlsx",
        help="Nombre del Excel de tarifa.",
    )
    args = parser.parse_args()

    base_dir = Path(args.base_dir)
    downloads = Path(args.downloads_dir)
    datos = base_dir / "datos_lims"

    sources = [
        (
            downloads / "PRISLAB_ReporteParametros.xlsx",
            {
                "Parametros": datos / "Parametros.csv",
                "Valores normalidad": datos / "Valores_normalidad.csv",
            },
        ),
        (
            downloads / "PRISLAB_ReporteExamenes.xlsx",
            {
                "Examenes": datos / "Examenes.csv",
                "Prueba del perfil": datos / "Examenes_Perfil.csv",
            },
        ),
        (
            downloads / "PRISLAB_ReportePaquete.xlsx",
            {
                "Paquetes": datos / "Paquetes.csv",
                "Prueba-perfil paquete": datos / "Paquetes_Perfil.csv",
            },
        ),
        (
            downloads / args.tarifa,
            {
                "Reporte": datos / "Tarifa_estudios de laboratorio.csv",
            },
        ),
    ]

    for workbook_path, sheets in sources:
        if not workbook_path.exists():
            raise FileNotFoundError(f"No existe el archivo origen: {workbook_path}")
        for sheet_name, output_path in sheets.items():
            _write_csv_from_sheet(workbook_path, sheet_name, output_path)
            print(f"[OK] {workbook_path.name} :: {sheet_name} -> {output_path.name}")


if __name__ == "__main__":
    main()
