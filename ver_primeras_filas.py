"""
Ver las primeras 5 filas del Excel para entender su estructura
"""
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from openpyxl import load_workbook

archivo = "Productos-farmacia-2026-02-10-10-31.xlsx"

print("=" * 80)
print("INSPECCION DEL EXCEL")
print("=" * 80)

wb = load_workbook(archivo, read_only=True, data_only=True)
ws = wb.active

print(f"Hoja: {ws.title}")
print(f"Filas: {ws.max_row}")
print(f"Columnas: {ws.max_column}")
print()

print("PRIMERAS 5 FILAS:")
print("=" * 80)

for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
    print(f"\nFILA {row_num}:")
    for col_num, cell in enumerate(row[:15], start=1):  # Primeras 15 columnas
        if cell:
            print(f"  Col {col_num}: {cell}")
