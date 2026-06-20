from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook

from core.models import Empresa, Medico


class Command(BaseCommand):
    help = "Importa catálogo de médicos desde el Excel legacy del laboratorio."

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str, help="Ruta al archivo Excel de médicos.")
        parser.add_argument("--empresa-id", type=int, default=None, help="Empresa destino.")
        parser.add_argument("--dry-run", action="store_true", help="Solo validar, sin guardar.")

    def handle(self, *args, **options):
        ruta = Path(options["archivo"]).expanduser()
        if not ruta.exists():
            raise CommandError(f"No existe el archivo: {ruta}")

        empresa = self._resolver_empresa(options.get("empresa_id"))
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        encabezado = None
        creados = 0
        actualizados = 0
        omitidos = 0

        for row in ws.iter_rows(values_only=True):
            values = [self._clean(v) for v in row]
            if not any(values):
                continue

            if encabezado is None and "Nombre" in values and "ID" in values:
                encabezado = {str(v).strip(): idx for idx, v in enumerate(values) if v is not None}
                continue

            if encabezado is None:
                continue

            legacy_id = values[encabezado["ID"]] if "ID" in encabezado else None
            nombre = values[encabezado["Nombre"]] if "Nombre" in encabezado else None
            especialidad = values[encabezado["Especialidad"]] if "Especialidad" in encabezado else None

            if not nombre:
                omitidos += 1
                continue

            cedula = self._build_cedula(legacy_id, nombre)
            defaults = {
                "empresa": empresa,
                "nombre_completo": str(nombre),
                "especialidad": str(especialidad or "Médico General"),
                "activo": True,
            }

            existente = Medico.objects.filter(cedula_profesional=cedula, empresa=empresa).first()
            if options["dry_run"]:
                if existente:
                    actualizados += 1
                else:
                    creados += 1
                continue

            medico, created = Medico.objects.update_or_create(
                cedula_profesional=cedula,
                empresa=empresa,
                defaults=defaults,
            )
            if created:
                creados += 1
            else:
                actualizados += 1

        self.stdout.write(self.style.SUCCESS("IMPORTACION MEDICOS COMPLETADA"))
        self.stdout.write(f"- empresa: {empresa.id} - {empresa.nombre}")
        self.stdout.write(f"- archivo: {ruta}")
        self.stdout.write(f"- creados: {creados}")
        self.stdout.write(f"- actualizados: {actualizados}")
        self.stdout.write(f"- omitidos: {omitidos}")
        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("dry-run activo: no se escribieron cambios"))

    def _resolver_empresa(self, empresa_id):
        if empresa_id:
            empresa = Empresa.objects.filter(pk=empresa_id).first()
            if not empresa:
                raise CommandError(f"No existe Empresa id={empresa_id}")
            return empresa
        empresa = Empresa.objects.filter(activa=True).order_by("pk").first()
        if not empresa:
            raise CommandError("No hay empresa activa disponible.")
        return empresa

    def _build_cedula(self, legacy_id, nombre):
        if legacy_id in (None, ""):
            return f"LEGACY-NO-ID-{str(nombre).strip().upper().replace(' ', '-')[:30]}"
        try:
            legacy_text = str(int(float(legacy_id)))
        except Exception:
            legacy_text = str(legacy_id).strip().upper().replace(" ", "-")
        return f"LEGACY-{legacy_text}"

    def _clean(self, value):
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            return value or None
        return value
