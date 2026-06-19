"""
CARGA MASIVA DE INVENTARIO DESDE XLSX (Excel)
==============================================
Lee el archivo Excel exportado del sistema POS y carga productos + lotes
manteniendo el ORDEN EXACTO del archivo original.

Columnas esperadas:
  1.  Nombre del Producto
  2.  Identificador (No Cambiar)
  3.  Es un Servicio
  4.  Se Vende
  5.  Descripción
  6.  Categoría
  7.  Marca
  8.  Unidad de Venta
  9.  Código de Barras
  10. SKU
  15. Usa Lotes
  16. Lote
  17. Fabricación del Lote
  18. Caducidad del Lote
  19. Utiliza Stock
  20. Stock Total
  21. Stock Mínimo
  22. Stock Apartado
  23. Ubicación
  24. Precio Público
  25. Precio PERSONAL
  26. Precio DESCUENTO PROMOCION
  27. Costo
  28. Impuestos
  29. IVA
  31. Clave SAT
  32. Receta Médica

Un mismo producto puede aparecer en MÚLTIPLES filas (una por lote).
El comando consolida stock total y crea los lotes individuales.
Puede detectar automáticamente el encabezado aunque el Excel traiga
filas introductorias antes de la tabla real.

Uso:
  python manage.py cargar_inventario_xlsx ruta/al/archivo.xlsx
  python manage.py cargar_inventario_xlsx ruta/al/archivo.xlsx --limpiar
"""
import os
import shutil
import tempfile
import zipfile
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from collections import OrderedDict
from xml.etree import ElementTree as ET

from django.core.management.base import BaseCommand
from django.db import transaction

from core.models import Producto, Lote, Sucursal
from core.utils.tenant_strict import add_argument_empresa_id, empresa_desde_management


class Command(BaseCommand):
    help = 'Carga masiva de productos de farmacia y lotes desde XLSX (Excel)'

    def add_arguments(self, parser):
        add_argument_empresa_id(parser, required=True)
        parser.add_argument('archivo', type=str, help='Ruta al archivo XLSX')
        parser.add_argument(
            '--limpiar', action='store_true',
            help='Eliminar TODOS los productos y lotes antes de cargar (reset total)'
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Solo simular sin guardar en la BD'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        limpiar = options['limpiar']
        dry_run = options['dry_run']

        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'Archivo no encontrado: {archivo}'))
            return

        try:
            import openpyxl
        except ImportError:
            self.stdout.write(self.style.ERROR('openpyxl no instalado. Ejecuta: pip install openpyxl'))
            return

        try:
            empresa = empresa_desde_management(options)
        except Exception as e:
            self.stdout.write(self.style.ERROR(str(e)))
            return

        sucursal = Sucursal.objects.filter(empresa=empresa).first()

        self.stdout.write('=' * 80)
        self.stdout.write(self.style.SUCCESS('CARGA MASIVA DE INVENTARIO DESDE XLSX'))
        self.stdout.write(f'  Archivo:  {archivo}')
        self.stdout.write(f'  Empresa:  {empresa.nombre}')
        self.stdout.write(f'  Sucursal: {sucursal.nombre if sucursal else "N/A"}')
        self.stdout.write(f'  Limpiar:  {"SÍ" if limpiar else "No"}')
        self.stdout.write(f'  Dry-run:  {"SÍ" if dry_run else "No"}')
        self.stdout.write('=' * 80)

        wb = self._abrir_workbook_seguro(openpyxl, archivo)
        ws = wb.active

        header_row_idx, col_map = self._detectar_encabezados(ws)
        if not col_map:
            self.stdout.write(self.style.ERROR(
                'No se detectó la fila de encabezados del inventario. '
                'Verifica que el archivo incluya "Nombre del Producto" e '
                '"Identificador (No Cambiar)".'
            ))
            wb.close()
            return
        self.stdout.write(f'  Columnas detectadas: {len(col_map)}')
        self.stdout.write(f'  Fila de encabezado detectada: {header_row_idx}')

        filas = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            nombre = row[col_map.get('Nombre del Producto', 0)]
            if nombre and str(nombre).strip():
                filas.append(row)
        wb.close()

        self.stdout.write(f'  Filas con datos: {len(filas)}')

        productos_ordenados = self._agrupar_preservando_orden(filas, col_map)
        self.stdout.write(f'  Productos únicos: {len(productos_ordenados)}')

        if dry_run:
            self.stdout.write(self.style.WARNING('\n  [DRY-RUN] No se guardará nada en la BD.'))
            for i, (clave, data) in enumerate(productos_ordenados.items(), 1):
                info = data['info']
                nombre = str(info[col_map['Nombre del Producto']]).strip()
                self.stdout.write(f'  {i:4d}. {nombre[:60]} | CB: {clave} | Stock: {data["stock_total"]} | Lotes: {len(data["lotes"])}')
                if i >= 30:
                    self.stdout.write(f'  ... y {len(productos_ordenados) - 30} más')
                    break
            return

        if limpiar:
            self.stdout.write(self.style.WARNING('\n  Limpiando productos y lotes existentes...'))
            from django.db.models.deletion import ProtectedError
            lotes_borrados = 0
            prods_borrados = 0
            prods_protegidos = 0
            for lote in Lote.objects.filter(producto__empresa=empresa).iterator():
                try:
                    lote.delete()
                    lotes_borrados += 1
                except ProtectedError:
                    pass
            for prod in Producto.objects.filter(empresa=empresa).iterator():
                try:
                    prod.delete()
                    prods_borrados += 1
                except ProtectedError:
                    prods_protegidos += 1
            self.stdout.write(f'  Eliminados: {prods_borrados} productos, {lotes_borrados} lotes')
            if prods_protegidos:
                self.stdout.write(f'  Protegidos (tienen ventas): {prods_protegidos} productos conservados')

        creados, actualizados, lotes_creados, errores = self._cargar_en_bd(
            productos_ordenados, col_map, empresa, sucursal
        )

        self.stdout.write('')
        self.stdout.write('=' * 80)
        self.stdout.write(self.style.SUCCESS('CARGA COMPLETADA'))
        self.stdout.write(f'  Productos creados:       {creados}')
        self.stdout.write(f'  Productos actualizados:  {actualizados}')
        self.stdout.write(f'  Lotes creados:           {lotes_creados}')
        if errores:
            self.stdout.write(self.style.WARNING(f'  Errores:                 {len(errores)}'))
            for e in errores[:20]:
                self.stdout.write(f'    - {e}')
            if len(errores) > 20:
                self.stdout.write(f'    ... y {len(errores) - 20} errores más')
        self.stdout.write('=' * 80)

    def _agrupar_preservando_orden(self, filas, col_map):
        """
        Agrupa filas por producto (código de barras o identificador),
        preservando el ORDEN de primera aparición en el archivo.
        """
        productos = OrderedDict()

        for fila in filas:
            cb_raw = fila[col_map.get('Código de Barras', 8)]
            ident = fila[col_map.get('Identificador (No Cambiar)', 1)]
            nombre = str(fila[col_map.get('Nombre del Producto', 0)]).strip()

            cb = str(cb_raw).strip() if cb_raw else ''
            ident = str(ident).strip() if ident else ''
            clave = cb or ident or f'AUTO-{abs(hash(nombre)) % 1000000:06d}'

            stock_raw = fila[col_map.get('Stock Total', 19)]
            try:
                stock = int(float(stock_raw)) if stock_raw is not None else 0
            except (ValueError, TypeError):
                stock = 0
            stock = max(stock, 0)

            if clave not in productos:
                productos[clave] = {
                    'info': fila,
                    'lotes': [],
                    'stock_total': 0,
                }

            productos[clave]['stock_total'] += stock

            lote_num_raw = fila[col_map.get('Lote', 15)]
            lote_num = str(lote_num_raw).strip() if lote_num_raw else ''
            caducidad_raw = fila[col_map.get('Caducidad del Lote', 17)]
            fabricacion_raw = fila[col_map.get('Fabricación del Lote', 16)]
            ubicacion_raw = fila[col_map.get('Ubicación', 22)]

            if lote_num:
                productos[clave]['lotes'].append({
                    'numero': lote_num,
                    'caducidad': caducidad_raw,
                    'fabricacion': fabricacion_raw,
                    'stock': stock,
                    'ubicacion': str(ubicacion_raw).strip() if ubicacion_raw else '',
                })

        return productos

    def _cargar_en_bd(self, productos_ordenados, col_map, empresa, sucursal):
        creados = 0
        actualizados = 0
        lotes_creados = 0
        errores = []

        with transaction.atomic():
            for idx, (clave, data) in enumerate(productos_ordenados.items(), 1):
                try:
                    info = data['info']
                    nombre = str(info[col_map['Nombre del Producto']]).strip()[:255]
                    if not nombre:
                        continue

                    cb_raw = info[col_map.get('Código de Barras', 8)]
                    codigo_barras = str(cb_raw).strip() if cb_raw else ''
                    if not codigo_barras:
                        ident = info[col_map.get('Identificador (No Cambiar)', 1)]
                        codigo_barras = str(ident).strip() if ident else f'SIN-CB-{abs(hash(nombre)) % 1000000:06d}'

                    marca_raw = info[col_map.get('Marca', 6)]
                    marca = str(marca_raw).strip()[:150] if marca_raw else 'GENÉRICO'

                    desc_raw = info[col_map.get('Descripción', 4)]
                    descripcion = str(desc_raw).strip()[:255] if desc_raw else ''

                    cat_raw = info[col_map.get('Categoría', 5)]
                    categoria = self._mapear_categoria(str(cat_raw).strip() if cat_raw else '')

                    precio_pub = self._parse_decimal(info[col_map.get('Precio Público', 23)])
                    precio_personal = self._parse_decimal(info[col_map.get('Precio PERSONAL', 24)])
                    costo = self._parse_decimal(info[col_map.get('Costo', 26)])

                    impuestos_raw = info[col_map.get('Impuestos', 27)]
                    iva_raw = info[col_map.get('IVA', 28)]
                    tiene_iva = str(impuestos_raw).strip().lower() in ('si', 'sí') if impuestos_raw else False
                    iva_pct = Decimal('16.00') if tiene_iva else Decimal('0.00')
                    if iva_raw and tiene_iva:
                        try:
                            iva_pct = Decimal(str(iva_raw)) * 100
                        except (InvalidOperation, ValueError):
                            iva_pct = Decimal('16.00')

                    es_servicio_raw = info[col_map.get('Es un Servicio', 2)]
                    es_servicio = str(es_servicio_raw).strip().lower() in ('si', 'sí') if es_servicio_raw else False

                    receta_raw = info[col_map.get('Receta Médica', 31)]
                    es_receta = str(receta_raw).strip().lower() in ('si', 'sí') if receta_raw else False

                    stock_total = data['stock_total']

                    stock_min_raw = info[col_map.get('Stock Mínimo', 20)]
                    try:
                        stock_minimo = int(float(stock_min_raw)) if stock_min_raw is not None else 0
                    except (ValueError, TypeError):
                        stock_minimo = 0

                    clasificacion = 'IV' if es_receta else 'VI'
                    if categoria == 'ANTIBIOTICO':
                        clasificacion = 'IV'

                    unidad_raw = info[col_map.get('Unidad de Venta', 7)]
                    unidad = str(unidad_raw).strip() if unidad_raw else 'Unidad'

                    producto, created = Producto.objects.update_or_create(
                        codigo_barras=codigo_barras,
                        defaults={
                            'empresa': empresa,
                            'sucursal': sucursal,
                            'nombre': nombre,
                            'sustancia_activa': descripcion if descripcion else '',
                            'marca_laboratorio': marca,
                            'forma_farmaceutica': unidad,
                            'concentracion': self._extraer_concentracion(nombre),
                            'presentacion': self._extraer_presentacion(nombre),
                            'precio_compra': costo,
                            'precio_publico': precio_pub,
                            'iva_porcentaje': iva_pct,
                            'stock': stock_total,
                            'stock_minimo': stock_minimo if stock_minimo > 0 else 0,
                            'clasificacion_sanitaria': clasificacion,
                            'categoria': categoria,
                            'es_antibiotico': es_receta or categoria == 'ANTIBIOTICO',
                            'es_servicio': es_servicio,
                        }
                    )

                    if created:
                        creados += 1
                    else:
                        actualizados += 1

                    for lote_data in data['lotes']:
                        fecha_cad = self._parse_fecha(lote_data['caducidad'])
                        fecha_fab = self._parse_fecha(lote_data['fabricacion'])

                        if not fecha_cad:
                            continue

                        lote_existente = Lote.objects.filter(
                            producto=producto,
                            numero_lote=lote_data['numero']
                        ).first()

                        if lote_existente:
                            lote_existente.cantidad = lote_data['stock']
                            lote_existente.fecha_caducidad = fecha_cad
                            if fecha_fab:
                                lote_existente.fecha_fabricacion = fecha_fab
                            if lote_data.get('ubicacion'):
                                lote_existente.ubicacion_fisica = lote_data['ubicacion'][:150]
                            Lote.objects.filter(pk=lote_existente.pk).update(
                                cantidad=lote_existente.cantidad,
                                fecha_caducidad=lote_existente.fecha_caducidad,
                                fecha_fabricacion=lote_existente.fecha_fabricacion,
                                ubicacion_fisica=lote_existente.ubicacion_fisica,
                            )
                        else:
                            Lote.objects.filter(pk=0).exists()
                            lote_obj = Lote(
                                producto=producto,
                                numero_lote=lote_data['numero'],
                                fecha_caducidad=fecha_cad,
                                fecha_fabricacion=fecha_fab,
                                cantidad=lote_data['stock'],
                                costo_adquisicion=costo if costo > 0 else Decimal('0.01'),
                                ubicacion_fisica=lote_data.get('ubicacion', '')[:150] if lote_data.get('ubicacion') else None,
                            )
                            super(Lote, lote_obj).save()
                            lotes_creados += 1

                    if idx % 100 == 0:
                        self.stdout.write(f'  Procesados: {idx}/{len(productos_ordenados)}...')

                except Exception as e:
                    errores.append(f'Fila {idx} ({clave}): {e}')

        return creados, actualizados, lotes_creados, errores

    def _abrir_workbook_seguro(self, openpyxl, archivo):
        """
        Abre el XLSX original y, si trae validaciones inválidas, genera una copia
        temporal sin nodos <dataValidations> para que openpyxl pueda leerlo.
        """
        if self._requiere_sanitizar_validaciones(archivo):
            copia_limpia = self._eliminar_validaciones_invalidas(archivo)
            self.stdout.write(self.style.WARNING(
                f'  Workbook sanitizado para importación segura: {copia_limpia}'
            ))
            return openpyxl.load_workbook(copia_limpia, read_only=True)

        try:
            return openpyxl.load_workbook(archivo, read_only=True)
        except ValueError as exc:
            mensaje = str(exc).lower()
            if 'value must be one of' not in mensaje and 'datavalidation' not in mensaje:
                raise

            copia_limpia = self._eliminar_validaciones_invalidas(archivo)
            self.stdout.write(self.style.WARNING(
                f'  Workbook sanitizado para abrirlo sin validaciones inválidas: {copia_limpia}'
            ))
            return openpyxl.load_workbook(copia_limpia, read_only=True)

    def _requiere_sanitizar_validaciones(self, archivo):
        try:
            with zipfile.ZipFile(archivo, 'r') as zin:
                for item in zin.namelist():
                    if not item.startswith('xl/worksheets/sheet') or not item.endswith('.xml'):
                        continue
                    data = zin.read(item)
                    if b'<dataValidations' in data:
                        return True
        except OSError:
            return False
        return False

    def _detectar_encabezados(self, ws, max_rows=10):
        """
        Algunos exports reales de PRISLAB traen 1-2 filas introductorias antes
        del encabezado. Escaneamos las primeras filas y usamos la primera que
        contenga las columnas clave del inventario.
        """
        requeridas = {'Nombre del Producto', 'Identificador (No Cambiar)'}

        for idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_rows, values_only=True),
            start=1,
        ):
            headers = [self._normalizar_header(cell) for cell in row]
            col_map = {header: pos for pos, header in enumerate(headers) if header}
            if requeridas.issubset(col_map.keys()):
                return idx, col_map

        return None, {}

    def _normalizar_header(self, value):
        if value is None:
            return ''
        return str(value).replace('\n', ' ').strip()

    def _eliminar_validaciones_invalidas(self, archivo):
        """
        Crea una copia temporal del XLSX eliminando los nodos de validación de datos
        que provocan errores de parseo en openpyxl.
        """
        tmp_dir = tempfile.mkdtemp(prefix='prislab_inv_')
        salida = os.path.join(tmp_dir, os.path.basename(archivo))

        ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        with zipfile.ZipFile(archivo, 'r') as zin, zipfile.ZipFile(salida, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/sheet') and item.filename.endswith('.xml'):
                    try:
                        root = ET.fromstring(data)
                        changed = False
                        for nodo in root.findall('main:dataValidations', ns):
                            root.remove(nodo)
                            changed = True
                        if changed:
                            data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                    except ET.ParseError:
                        pass
                zout.writestr(item, data)

        return salida

    def _parse_decimal(self, valor):
        if valor is None:
            return Decimal('0.00')
        try:
            limpio = str(valor).replace('$', '').replace(',', '').strip()
            return Decimal(limpio).quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError):
            return Decimal('0.00')

    def _parse_fecha(self, valor):
        if valor is None:
            return None
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        fecha_str = str(valor).strip()
        if not fecha_str:
            return None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(fecha_str, fmt).date()
            except ValueError:
                continue
        return None

    def _mapear_categoria(self, cat_raw):
        if not cat_raw:
            return 'GENERICO'
        cat = cat_raw.lower()
        if 'antibi' in cat:
            return 'ANTIBIOTICO'
        elif 'patente' in cat:
            return 'PATENTE'
        elif 'curacion' in cat or 'material' in cat:
            return 'CURACION'
        elif 'medicamento' in cat:
            return 'GENERICO'
        elif 'generico' in cat or 'genérico' in cat:
            return 'GENERICO'
        return 'OTRO'

    def _extraer_concentracion(self, nombre):
        """Intenta extraer concentración del nombre (ej: 500MG, 1G/2ML)."""
        import re
        match = re.search(r'(\d+(?:\.\d+)?(?:MG|G|ML|MCG|UI|U|MG/ML|G/ML|MG/\d+ML|G/\d+ML)(?:/\d+(?:MG|G|ML))?)', nombre.upper())
        if match:
            return match.group(1)
        return 'N/A'

    def _extraer_presentacion(self, nombre):
        """Intenta extraer presentación del nombre (ej: (30), (10))."""
        import re
        match = re.search(r'\((\d+)\)', nombre)
        if match:
            return match.group(1)
        return '1'
