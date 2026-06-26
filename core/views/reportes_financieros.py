"""
Módulo de Reportes Financieros Detallados - PRISLAB
Reportes financieros avanzados: P&L, Balance, Flujo de Caja, etc.
"""
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from core.decorators import role_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden
from django.db.models import Sum, Count, Q, F
from django.db.models.functions import Coalesce, TruncMonth, TruncDay
from django.db.models import DecimalField
from django.utils import timezone
from decimal import Decimal
from datetime import datetime, timedelta
import json

from core.models import (
    Venta, Pago, GastoCaja, GastoOperativo
    # NOTA: Modelos Compra, PolizaContable, MovimientoContable, CatalogoCuenta y Nomina pendientes de migración.
    # Compra, PolizaContable, MovimientoContable, CatalogoCuenta, Nomina
)


def _sumas_por_dia(queryset, fecha_field: str, total_field: str):
    """
    Agrupa montos por día y retorna un dict {date: Decimal}.
    Reduce loops N+1 en reportes diarios.
    """
    return {
        item['dia'].date(): item['total'] or Decimal('0.00')
        for item in queryset.annotate(dia=TruncDay(fecha_field))
        .values('dia')
        .annotate(total=Coalesce(Sum(total_field), Decimal('0.00'), output_field=DecimalField()))
        .order_by('dia')
    }


@login_required
@role_required('DIRECTOR', 'ADMIN', 'GERENTE', 'FINANZAS')
def reporte_ingresos_egresos(request):
    """Reporte de Ingresos y Egresos (P&L simplificado)."""
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario sin empresa asignada.')
        return redirect('home')

    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Por defecto, último mes
    hoy = timezone.localdate()
    if not fecha_inicio:
        fecha_inicio = (hoy - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_fin:
        fecha_fin = hoy.strftime('%Y-%m-%d')
    
    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    # INGRESOS
    ventas = Venta.objects.filter(
        empresa=empresa,
        fecha__date__range=[fecha_inicio_dt, fecha_fin_dt],
        estado='COMPLETADA'
    )
    total_ventas = ventas.aggregate(
        total=Coalesce(Sum('total'), Decimal('0.00'), output_field=DecimalField())
    )['total'] or Decimal('0.00')
    
    # EGRESOS (Compra no migrado aún)
    total_compras = Decimal('0.00')

    gastos_caja = GastoCaja.objects.filter(
        empresa=empresa,
        fecha__date__range=[fecha_inicio_dt, fecha_fin_dt]
    )
    total_gastos_caja = gastos_caja.aggregate(
        total=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField())
    )['total'] or Decimal('0.00')

    gastos_operativos = GastoOperativo.objects.filter(
        empresa=empresa,
        fecha__date__range=[fecha_inicio_dt, fecha_fin_dt]
    )
    total_gastos_operativos = gastos_operativos.aggregate(
        total=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField())
    )['total'] or Decimal('0.00')

    # Nóminas pagadas (Nomina no migrado aún)
    total_nominas = Decimal('0.00')
    
    total_egresos = total_compras + total_gastos_caja + total_gastos_operativos + total_nominas
    utilidad_bruta = total_ventas - total_compras
    utilidad_neta = total_ventas - total_egresos
    
    # Datos para gráfica
    ventas_por_dia = _sumas_por_dia(ventas, 'fecha', 'total')
    gastos_caja_por_dia = _sumas_por_dia(gastos_caja, 'fecha', 'monto')
    gastos_operativos_por_dia = _sumas_por_dia(gastos_operativos, 'fecha', 'monto')

    datos_diarios = []
    fecha_actual = fecha_inicio_dt
    while fecha_actual <= fecha_fin_dt:
        ventas_dia = ventas_por_dia.get(fecha_actual, Decimal('0.00'))
        egresos_caja_dia = gastos_caja_por_dia.get(fecha_actual, Decimal('0.00'))
        egresos_operativos_dia = gastos_operativos_por_dia.get(fecha_actual, Decimal('0.00'))
        egresos_dia = egresos_caja_dia + egresos_operativos_dia
        
        datos_diarios.append({
            'fecha': fecha_actual.strftime('%Y-%m-%d'),
            'ingresos': float(ventas_dia),
            'egresos': float(egresos_dia),
            'utilidad': float(ventas_dia - egresos_dia),
        })
        fecha_actual += timedelta(days=1)
    
    return render(request, 'core/reportes_financieros/ingresos_egresos.html', {
        'empresa': empresa,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'total_ventas': total_ventas,
        'total_compras': total_compras,
        'total_gastos_caja': total_gastos_caja,
        'total_gastos_operativos': total_gastos_operativos,
        'total_nominas': total_nominas,
        'total_egresos': total_egresos,
        'utilidad_bruta': utilidad_bruta,
        'utilidad_neta': utilidad_neta,
        'datos_diarios': json.dumps(datos_diarios),
    })


@login_required
@role_required('DIRECTOR', 'ADMIN', 'GERENTE', 'FINANZAS')
def reporte_balance_general(request):
    """Reporte de Balance General (Activos, Pasivos, Capital)."""
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario sin empresa asignada.')
        return redirect('home')

    # Fecha de corte
    fecha_corte = request.GET.get('fecha_corte', timezone.localdate().strftime('%Y-%m-%d'))
    fecha_corte_dt = datetime.strptime(fecha_corte, '%Y-%m-%d').date()

    from contabilidad.models import CuentaContable, AsientoContable

    saldos = {}
    for asiento in AsientoContable.objects.filter(
        poliza__empresa=empresa,
        poliza__fecha__lte=fecha_corte_dt,
        poliza__estado='AUTORIZADA',
    ).select_related('cuenta'):
        saldos.setdefault(asiento.cuenta_id, {'cuenta': asiento.cuenta, 'cargo': Decimal('0'), 'abono': Decimal('0')})
        saldos[asiento.cuenta_id]['cargo'] += asiento.cargo
        saldos[asiento.cuenta_id]['abono'] += asiento.abono

    def _saldo(cuenta, cargo, abono):
        return (cargo - abono) if cuenta.naturaleza == 'DEUDOR' else (abono - cargo)

    activos, pasivos, capital = [], [], []
    total_activos = Decimal('0')
    total_pasivos = Decimal('0')
    total_capital = Decimal('0')
    total_ingresos = Decimal('0')
    total_gastos = Decimal('0')
    total_costos = Decimal('0')

    for cuenta in CuentaContable.objects.filter(empresa=empresa, activa=True):
        s = saldos.get(cuenta.id, {'cargo': Decimal('0'), 'abono': Decimal('0')})
        saldo = _saldo(cuenta, s['cargo'], s['abono'])
        if saldo == 0:
            continue
        item = {'codigo': cuenta.codigo, 'nombre': cuenta.nombre, 'saldo': saldo}
        if cuenta.tipo == 'ACTIVO':
            activos.append(item); total_activos += saldo
        elif cuenta.tipo == 'PASIVO':
            pasivos.append(item); total_pasivos += saldo
        elif cuenta.tipo == 'CAPITAL':
            capital.append(item); total_capital += saldo
        elif cuenta.tipo == 'INGRESO':
            total_ingresos += saldo
        elif cuenta.tipo == 'GASTO':
            total_gastos += saldo
        elif cuenta.tipo == 'COSTO':
            total_costos += saldo

    hay_asientos = bool(saldos)
    if not hay_asientos:
        messages.info(
            request,
            'No existen asientos contables autorizados para este período. '
            'Capture y autorice pólizas para obtener un balance real.'
        )

    utilidad = total_ingresos - total_gastos - total_costos
    total_capital += utilidad

    return render(request, 'core/reportes_financieros/balance_general.html', {
        'empresa': empresa,
        'fecha_corte': fecha_corte,
        'activos': activos,
        'total_activos': total_activos,
        'pasivos': pasivos,
        'total_pasivos': total_pasivos,
        'capital': capital,
        'total_capital': total_capital,
        'total_ingresos': total_ingresos,
        'total_gastos': total_gastos,
        'total_costos': total_costos,
        'utilidad': utilidad,
        'hay_asientos': hay_asientos,
    })


@login_required
@role_required('DIRECTOR', 'ADMIN', 'GERENTE', 'FINANZAS')
def reporte_flujo_caja(request):
    """Reporte de Flujo de Caja."""
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario sin empresa asignada.')
        return redirect('home')

    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    hoy = timezone.localdate()
    if not fecha_inicio:
        fecha_inicio = (hoy - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_fin:
        fecha_fin = hoy.strftime('%Y-%m-%d')
    
    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    # ENTRADAS DE EFECTIVO
    pagos_efectivo = Pago.objects.filter(
        venta__empresa=empresa,
        venta__fecha__date__range=[fecha_inicio_dt, fecha_fin_dt],
        venta__estado='COMPLETADA',
        metodo='EFECTIVO'
    )
    total_entradas_efectivo = pagos_efectivo.aggregate(
        total=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField())
    )['total'] or Decimal('0.00')
    
    # SALIDAS DE EFECTIVO
    gastos_caja = GastoCaja.objects.filter(
        empresa=empresa,
        fecha__date__range=[fecha_inicio_dt, fecha_fin_dt]
    )
    total_salidas_efectivo = gastos_caja.aggregate(
        total=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField())
    )['total'] or Decimal('0.00')
    
    # FLUJO NETO
    flujo_neto = total_entradas_efectivo - total_salidas_efectivo
    
    # Detalle diario
    pagos_por_dia = _sumas_por_dia(pagos_efectivo, 'fecha_pago', 'monto')
    gastos_caja_por_dia = _sumas_por_dia(gastos_caja, 'fecha', 'monto')

    flujo_diario = []
    fecha_actual = fecha_inicio_dt
    while fecha_actual <= fecha_fin_dt:
        entradas_dia = pagos_por_dia.get(fecha_actual, Decimal('0.00'))
        salidas_dia = gastos_caja_por_dia.get(fecha_actual, Decimal('0.00'))
        
        flujo_diario.append({
            'fecha': fecha_actual.strftime('%Y-%m-%d'),
            'entradas': float(entradas_dia),
            'salidas': float(salidas_dia),
            'flujo_neto': float(entradas_dia - salidas_dia),
        })
        fecha_actual += timedelta(days=1)
    
    return render(request, 'core/reportes_financieros/flujo_caja.html', {
        'empresa': empresa,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'total_entradas_efectivo': total_entradas_efectivo,
        'total_salidas_efectivo': total_salidas_efectivo,
        'flujo_neto': flujo_neto,
        'flujo_diario': json.dumps(flujo_diario),
    })


@login_required
@role_required('DIRECTOR', 'ADMIN', 'GERENTE', 'FINANZAS')
def api_ventas_por_mes(request):
    """API para obtener ventas agrupadas por mes."""
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        return JsonResponse(
            {'error': 'Usuario sin empresa asignada.'},
            status=403,
        )
    try:
        anio = int(request.GET.get('anio', timezone.localdate().year))
    except (TypeError, ValueError):
        anio = timezone.localdate().year

    ventas = Venta.objects.filter(
        empresa=empresa,
        fecha__year=anio,
        estado='COMPLETADA'
    ).annotate(
        mes=TruncMonth('fecha')
    ).values('mes').annotate(
        total=Coalesce(Sum('total'), Decimal('0.00'), output_field=DecimalField())
    ).order_by('mes')
    
    datos = {
        'labels': [],
        'valores': [],
    }
    
    for venta in ventas:
        datos['labels'].append(venta['mes'].strftime('%B'))
        datos['valores'].append(float(venta['total']))
    
    return JsonResponse(datos)


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIONES EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _excel_response(wb, filename):
    """Helper: convierte un workbook de openpyxl a HttpResponse."""
    from django.http import HttpResponse
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _estilo_encabezado(ws, row, cols, fill_hex='1F4E79'):
    """Aplica estilo de encabezado a una fila."""
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


@login_required
@role_required('DIRECTOR', 'ADMIN', 'GERENTE', 'FINANZAS')
def exportar_excel_ingresos_egresos(request):
    """Exporta el reporte de Ingresos y Egresos a Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        return HttpResponseForbidden('Usuario sin empresa asignada.')

    hoy = timezone.localdate()
    fecha_inicio = request.GET.get('fecha_inicio', (hoy - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', hoy.strftime('%Y-%m-%d'))
    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ingresos y Egresos'

    # Encabezado empresa
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte de Ingresos y Egresos — {empresa.nombre if empresa else ''}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Período: {fecha_inicio} al {fecha_fin}"
    ws['A2'].font = Font(italic=True, color='666666')
    ws.append([])

    # Resumen financiero
    ws.append(['CONCEPTO', '', 'MONTO (MXN)'])
    _estilo_encabezado(ws, ws.max_row, 3)

    # Calcular datos
    ventas_qs = Venta.objects.filter(
        empresa=empresa,
        fecha__date__range=[fecha_inicio_dt, fecha_fin_dt],
        estado='COMPLETADA'
    )
    total_ventas = ventas_qs.aggregate(
        t=Coalesce(Sum('total'), Decimal('0.00'), output_field=DecimalField())
    )['t'] or Decimal('0.00')

    gastos_caja_total = GastoCaja.objects.filter(
        empresa=empresa,
        fecha__date__range=[fecha_inicio_dt, fecha_fin_dt]
    ).aggregate(t=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField()))['t'] or Decimal('0.00')

    gastos_op_total = GastoOperativo.objects.filter(
        empresa=empresa,
        fecha__date__range=[fecha_inicio_dt, fecha_fin_dt]
    ).aggregate(t=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField()))['t'] or Decimal('0.00')

    total_egresos = gastos_caja_total + gastos_op_total
    utilidad = total_ventas - total_egresos

    ws.append(['INGRESOS', '', float(total_ventas)])
    ws.append(['  Ventas totales', '', float(total_ventas)])
    ws.append(['EGRESOS', '', float(total_egresos)])
    ws.append(['  Gastos de caja', '', float(gastos_caja_total)])
    ws.append(['  Gastos operativos', '', float(gastos_op_total)])
    ws.append([])
    ws.append(['UTILIDAD NETA', '', float(utilidad)])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.cell(ws.max_row, 3).font = Font(bold=True, color='006100' if utilidad >= 0 else 'C00000')

    ws.append([])
    ws.append([])

    # Detalle diario
    ws.append(['DETALLE DIARIO'])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.append(['Fecha', 'Ingresos', 'Egresos', 'Utilidad del día'])
    _estilo_encabezado(ws, ws.max_row, 4)

    ventas_por_dia = _sumas_por_dia(ventas_qs, 'fecha', 'total')
    gastos_caja_por_dia = _sumas_por_dia(
        GastoCaja.objects.filter(
            empresa=empresa,
            fecha__date__range=[fecha_inicio_dt, fecha_fin_dt]
        ),
        'fecha',
        'monto',
    )
    gastos_op_por_dia = _sumas_por_dia(
        GastoOperativo.objects.filter(
            empresa=empresa,
            fecha__date__range=[fecha_inicio_dt, fecha_fin_dt]
        ),
        'fecha',
        'monto',
    )

    fecha_actual = fecha_inicio_dt
    while fecha_actual <= fecha_fin_dt:
        ing = ventas_por_dia.get(fecha_actual, Decimal('0.00'))
        eg_caja = gastos_caja_por_dia.get(fecha_actual, Decimal('0.00'))
        eg_operativos = gastos_op_por_dia.get(fecha_actual, Decimal('0.00'))
        eg = eg_caja + eg_operativos
        ws.append([fecha_actual.strftime('%d/%m/%Y'), float(ing), float(eg), float(ing - eg)])
        fecha_actual += timedelta(days=1)

    # Ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    return _excel_response(wb, f'ingresos_egresos_{fecha_inicio}_{fecha_fin}.xlsx')


@login_required
@role_required('DIRECTOR', 'ADMIN', 'GERENTE', 'FINANZAS')
def exportar_excel_flujo_caja(request):
    """Exporta el reporte de Flujo de Caja a Excel."""
    import openpyxl
    from openpyxl.styles import Font

    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        return HttpResponseForbidden('Usuario sin empresa asignada.')

    hoy = timezone.localdate()
    fecha_inicio = request.GET.get('fecha_inicio', (hoy - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', hoy.strftime('%Y-%m-%d'))
    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Flujo de Caja'

    ws.merge_cells('A1:D1')
    ws['A1'] = f"Flujo de Caja — {empresa.nombre if empresa else ''}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Período: {fecha_inicio} al {fecha_fin}"
    ws.append([])
    ws.append(['Fecha', 'Entradas (Efectivo)', 'Salidas (Efectivo)', 'Flujo Neto'])
    _estilo_encabezado(ws, ws.max_row, 4)

    total_entradas = Decimal('0.00')
    total_salidas = Decimal('0.00')

    pagos_efectivo = Pago.objects.filter(
        venta__empresa=empresa,
        venta__fecha__date__range=[fecha_inicio_dt, fecha_fin_dt],
        venta__estado='COMPLETADA',
        metodo='EFECTIVO'
    )
    gastos_caja = GastoCaja.objects.filter(
        empresa=empresa,
        fecha__date__range=[fecha_inicio_dt, fecha_fin_dt]
    )

    entradas_por_dia = _sumas_por_dia(pagos_efectivo, 'fecha_pago', 'monto')
    salidas_por_dia = _sumas_por_dia(gastos_caja, 'fecha', 'monto')

    fecha_actual = fecha_inicio_dt
    while fecha_actual <= fecha_fin_dt:
        entradas = entradas_por_dia.get(fecha_actual, Decimal('0.00'))
        salidas = salidas_por_dia.get(fecha_actual, Decimal('0.00'))

        ws.append([
            fecha_actual.strftime('%d/%m/%Y'),
            float(entradas), float(salidas), float(entradas - salidas)
        ])
        total_entradas += entradas
        total_salidas += salidas
        fecha_actual += timedelta(days=1)

    ws.append([])
    ws.append(['TOTALES', float(total_entradas), float(total_salidas), float(total_entradas - total_salidas)])
    ws.cell(ws.max_row, 1).font = Font(bold=True)

    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 22

    return _excel_response(wb, f'flujo_caja_{fecha_inicio}_{fecha_fin}.xlsx')


@login_required
@role_required('DIRECTOR', 'ADMIN', 'GERENTE', 'FINANZAS')
def exportar_excel_balance(request):
    """Exporta el Balance General a Excel."""
    import openpyxl
    from openpyxl.styles import Font

    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        return HttpResponseForbidden('Usuario sin empresa asignada.')

    fecha_corte = request.GET.get('fecha_corte', timezone.localdate().strftime('%Y-%m-%d'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Balance General'

    ws.merge_cells('A1:C1')
    ws['A1'] = f"Balance General — {empresa.nombre if empresa else ''}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Fecha de corte: {fecha_corte}"
    ws.append([])

    # Activos — desde ventas/pagos como proxy si no hay contabilidad completa
    total_cxc = Pago.objects.filter(
        venta__empresa=empresa,
        venta__estado='COMPLETADA'
    ).aggregate(t=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField()))['t'] or Decimal('0.00')

    total_gastos = GastoCaja.objects.filter(empresa=empresa).aggregate(
        t=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField())
    )['t'] or Decimal('0.00')

    ws.append(['ACTIVOS', '', ''])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    _estilo_encabezado(ws, ws.max_row, 3, '1F4E79')
    ws.append(['  Cuentas por cobrar (ventas cobradas)', '', float(total_cxc)])
    ws.append(['Total Activos', '', float(total_cxc)])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.append([])

    ws.append(['PASIVOS / EGRESOS', '', ''])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    _estilo_encabezado(ws, ws.max_row, 3, 'C00000')
    ws.append(['  Gastos de caja acumulados', '', float(total_gastos)])
    ws.append(['Total Pasivos', '', float(total_gastos)])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.append([])

    capital = total_cxc - total_gastos
    ws.append(['CAPITAL / PATRIMONIO', '', float(capital)])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.cell(ws.max_row, 3).font = Font(bold=True, color='006100' if capital >= 0 else 'C00000')

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['C'].width = 20

    return _excel_response(wb, f'balance_general_{fecha_corte}.xlsx')
