"""
Script de Emergencia: Carga Excel con errores
Intenta leer fila por fila ignorando corrupciones
"""
import os
import sys
import django
import logging

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from core.models import Producto, Empresa, Sucursal

_eid = os.environ.get("PRISLAB_EMPRESA_ID")
if not _eid:
    print("[ERROR] Defina PRISLAB_EMPRESA_ID (pk de Empresa).")
    sys.exit(1)
try:
    _empresa_ref = Empresa.objects.get(pk=int(_eid))
except (ValueError, Empresa.DoesNotExist):
    print(f"[ERROR] Empresa id={_eid!r} no válida.")
    sys.exit(1)
from decimal import Decimal
from openpyxl import load_workbook

archivo = "Productos-farmacia-2026-02-10-10-31.xlsx"

print("=" * 80)
print("CARGA FORZADA DE INVENTARIO")
print("=" * 80)

try:
    # Intentar abrir con read_only y data_only
    wb = load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active
    
    print(f"[OK] Excel abierto: {ws.title}")
    print(f"[OK] Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
    print()
    
    empresa = _empresa_ref
    sucursal = Sucursal.objects.filter(empresa=empresa).first()
    
    # Leer headers (están en la fila 3, no en la 1)
    headers = [cell.value for cell in ws[3]]
    print("[COLUMNAS DETECTADAS]:")
    for i, h in enumerate(headers[:15], 1):
        if h:
            print(f"  {i}. {h}")
    print()
    
    # Mapear columnas
    col_map = {}
    for idx, h in enumerate(headers):
        if not h:
            continue
        h_str = str(h).strip()
        if 'Nombre del Producto' in h_str:
            col_map['nombre'] = idx
        elif 'Código de Barras' in h_str:
            col_map['codigo'] = idx
        elif 'Marca' in h_str:
            col_map['marca'] = idx
        elif 'Precio Público' in h_str:
            col_map['precio'] = idx
        elif 'Costo' in h_str:
            col_map['costo'] = idx
        elif 'Stock Total' in h_str:
            col_map['stock'] = idx
        elif 'IVA' in h_str:
            col_map['iva'] = idx
        elif 'Receta Médica' in h_str:
            col_map['receta'] = idx
    
    print("[MAPEO]:")
    for k, v in col_map.items():
        print(f"  {k}: Columna {v+1}")
    print()
    
    if 'nombre' not in col_map or 'codigo' not in col_map:
        print("[ERROR] No se encontraron columnas críticas")
        exit(1)
    
    # Procesar filas (datos empiezan en fila 4)
    print("[PROCESANDO...]")
    creados = 0
    errores = 0
    actualizados = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        try:
            nombre = row[col_map['nombre']] if col_map['nombre'] < len(row) else None
            codigo = row[col_map['codigo']] if col_map['codigo'] < len(row) else None
            
            if not nombre or not codigo:
                continue
            
            marca = row[col_map.get('marca', 0)] if col_map.get('marca') else 'GENERICO'
            
            # Precio
            precio = Decimal('0')
            if col_map.get('precio'):
                try:
                    val = row[col_map['precio']]
                    if val:
                        precio = Decimal(str(val).replace(',', '').replace('$', '').strip())
                except:
                    pass
            
            # Costo
            costo = Decimal('0')
            if col_map.get('costo'):
                try:
                    val = row[col_map['costo']]
                    if val:
                        costo = Decimal(str(val).replace(',', '').replace('$', '').strip())
                except:
                    pass
            
            # Stock
            stock = 0
            if col_map.get('stock'):
                try:
                    val = row[col_map['stock']]
                    if val:
                        stock = int(float(val))
                except:
                    pass
            
            # IVA
            iva = Decimal('16')
            if col_map.get('iva'):
                try:
                    val = str(row[col_map['iva']]).replace('%', '').strip()
                    if val and val != 'None':
                        iva = Decimal(val)
                except:
                    pass
            
            # Receta
            es_antibiotico = False
            if col_map.get('receta'):
                val = str(row[col_map['receta']]).lower()
                es_antibiotico = 'sí' in val or 'si' in val or 'obligatorio' in val
            
            # Crear producto
            producto, created = Producto.objects.update_or_create(
                codigo_barras=str(codigo),
                defaults={
                    'empresa': empresa,
                    'sucursal': sucursal,
                    'nombre': str(nombre)[:255],
                    'marca_laboratorio': str(marca)[:150] if marca else 'GENERICO',
                    'forma_farmaceutica': 'Pieza',
                    'concentracion': 'N/A',
                    'presentacion': '1',
                    'precio_compra': costo,
                    'precio_publico': precio,
                    'stock': stock,
                    'iva_porcentaje': iva,
                    'clasificacion_sanitaria': 'IV' if es_antibiotico else 'VI',
                    'categoria': 'ANTIBIOTICO' if es_antibiotico else 'GENERICO',
                    'es_antibiotico': es_antibiotico,
                    'es_servicio': False,
                }
            )
            
            if created:
                creados += 1
            else:
                actualizados += 1
            
            if (creados + actualizados) % 100 == 0:
                print(f"  [{creados + actualizados} productos procesados...]")
        
        except Exception as e:
            logging.getLogger(__name__).exception("Error inesperado en funcion_desconocida (cargar_excel_forzado.py)")
            errores += 1
            if errores <= 5:
                print(f"  [!] Error fila {row_idx}: {e}")
    
    print()
    print("=" * 80)
    print("[COMPLETADO]")
    print("=" * 80)
    print(f"Productos NUEVOS: {creados}")
    print(f"Productos ACTUALIZADOS: {actualizados}")
    print(f"Errores: {errores}")
    print(f"Total en sistema: {Producto.objects.count()}")
    print("=" * 80)
    
except Exception as e:
    logging.getLogger(__name__).exception("Error inesperado en funcion_desconocida (cargar_excel_forzado.py)")
    print(f"[ERROR CRITICO] {e}")
    import traceback
    traceback.print_exc()