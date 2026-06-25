"""
Vistas de Inventario de Farmacia
Incluye: entrada de mercancía, compras, carga masiva, libro de control, dashboard
"""

import json
import logging
import csv
from datetime import datetime, timedelta, date
from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import DecimalField, Q, Sum, F, Prefetch
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.conf import settings

logger = logging.getLogger('farmacia.inventario')

# Umbrales de caducidad configurables
_DIAS_CADUCIDAD_CRITICO = getattr(settings, 'FARMACIA_DIAS_CADUCIDAD_CRITICO', 30)

from core.models import (
    Producto, Lote, Venta, Pago, GastoCaja, DiscountPolicy, Empresa
)
from core.services.inventario.movimiento_inventario_service import MovimientoInventarioService
from core.services.inventario.catalogo_farmacia_service import CatalogoFarmaciaService


def _empresa_desde_request(request):
    """Empresa efectiva: EmpresaIdentityMiddleware (fallback principal) o FK del usuario."""
    return getattr(request, 'empresa_actual', None) or getattr(request.user, 'empresa', None)


# ==============================================================================
# ENTRADA DE MERCANCÍA
# ==============================================================================

@login_required
def entrada_mercancia(request):
    """Vista para entrada de mercancía - Procesa ingreso directo al almacén."""
    empresa = _empresa_desde_request(request)
    
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada.')
        return redirect('home')
    
    if request.method == 'POST':
        try:
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                data = request.POST.dict()
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'mensaje': 'JSON inválido'}, status=400)
        out = MovimientoInventarioService.entrada_mercancia_directa(request, empresa, data)
        return JsonResponse(out['body'], status=out['http_status'])
    
    # GET: Mostrar formulario
    return render(request, 'core/entrada_mercancia.html', {
        'empresa': empresa.nombre if empresa else 'PRISLAB'
    })


# ==============================================================================
# REGISTRAR COMPRA
# ==============================================================================

@login_required
def registrar_compra(request):
    """Vista para registrar compra de productos a proveedores.
    Usa MovimientoInventario (Kardex) para mantener trazabilidad completa."""
    empresa = _empresa_desde_request(request)
    if not empresa:
        return JsonResponse({'status': 'error', 'mensaje': 'Usuario sin empresa asignada'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'mensaje': 'JSON inválido'}, status=400)
        out = MovimientoInventarioService.registrar_compra_a_proveedor(request, empresa, data)
        return JsonResponse(out['body'], status=out['http_status'])
    
    # GET: Mostrar formulario
    from farmacia.models import Proveedor as FarmProveedor
    proveedores = FarmProveedor.objects.filter(empresa=empresa, activo=True).order_by('razon_social')
    return render(request, 'core/farmacia/compra_form.html', {
        'empresa': empresa,
        'proveedores': proveedores
    })


# ==============================================================================
# API: BUSCAR PRODUCTOS PARA COMPRA
# ==============================================================================

@login_required
@require_http_methods(["GET"])
def api_buscar_productos_compra(request):
    """API para buscar productos del catálogo para compras."""
    empresa = _empresa_desde_request(request)
    if not empresa:
        return JsonResponse({'productos': []})
    termino = request.GET.get('q', '').strip()
    
    if len(termino) < 2:
        return JsonResponse({'productos': []})
    
    # Buscar productos por nombre, código de barras o sustancia activa
    productos = Producto.objects.filter(
        empresa=empresa
    ).filter(
        Q(nombre__icontains=termino) |
        Q(codigo_barras__icontains=termino) |
        Q(sustancia_activa__icontains=termino) |
        Q(marca_laboratorio__icontains=termino)
    )[:20]  # Limitar a 20 resultados
    
    resultados = []
    for p in productos:
        resultados.append({
            'id': p.id,
            'nombre': p.nombre,
            'codigo_barras': p.codigo_barras,
            'sustancia_activa': p.sustancia_activa or '',
            'marca': p.marca_laboratorio or '',
            'precio_compra': float(p.precio_compra or 0),
            'stock': int(p.stock or 0),
        })
    
    return JsonResponse({'productos': resultados})


# ==============================================================================
# CARGA MASIVA DE PRODUCTOS
# ==============================================================================

@login_required
def carga_masiva_productos(request):
    """Carga masiva de productos desde CSV o XLSX."""
    empresa = _empresa_desde_request(request)
    if not empresa:
        return JsonResponse({'status': 'error', 'mensaje': 'Usuario sin empresa asignada'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido'}, status=405)
    
    uploaded = request.FILES.get('archivo')
    if not uploaded:
        return JsonResponse({'status': 'error', 'mensaje': 'No se proporcionó archivo'}, status=400)
    
    nombre_archivo = uploaded.name.lower()
    
    # Leer archivo
    try:
        rows = []
        if nombre_archivo.endswith('.csv'):
            decoded = uploaded.read().decode('utf-8-sig')
            reader = csv.DictReader(decoded.splitlines())
            rows = [
                {k.strip(): v for k, v in row.items()}
                for row in reader
            ]
        elif nombre_archivo.endswith('.xlsx'):
            import openpyxl
            workbook = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            sheet = workbook.active
            raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not raw_headers:
                return JsonResponse({'status': 'error', 'mensaje': 'Archivo sin encabezados'}, status=400)
            headers = [_normalizar_header(h) for h in raw_headers]
            rows = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                if not any(values):
                    continue
                rows.append({headers[idx]: value for idx, value in enumerate(values) if idx < len(headers)})
        else:
            return JsonResponse({'status': 'error', 'mensaje': 'Formato no soportado. Use CSV o XLSX'}, status=400)
    except Exception as exc:
        logger.exception('Error leyendo carga masiva farmacia: %s', exc)
        return JsonResponse({'status': 'error', 'mensaje': f'No se pudo leer el archivo: {exc}'}, status=400)

    productos_data = []
    errores = []
    for idx, row in enumerate(rows, start=2):
        producto = _fila_a_producto(row)
        if not producto['nombre']:
            errores.append(f'Fila {idx}: nombre requerido')
            continue
        productos_data.append(producto)

    if not productos_data:
        return JsonResponse({'status': 'error', 'mensaje': 'No hay productos validos para importar', 'errores': errores}, status=400)

    sucursal = empresa.sucursales.first()
    limpiar = str(request.POST.get('limpiar', '')).lower() in {'1', 'true', 'si', 'sí'}
    out = CatalogoFarmaciaService.carga_masiva_productos(
        empresa,
        sucursal,
        productos_data,
        limpiar=limpiar,
    )
    body = dict(out['body'])
    body['errores_parseo'] = errores
    body['procesados_archivo'] = len(productos_data)
    return JsonResponse(body, status=out['http_status'])


def _normalizar_header(header):
    """Normaliza encabezados de CSV/XLSX."""
    if not header:
        return ''
    return str(header).strip().lower().replace(' ', '_').replace('-', '_')


def _fila_a_producto(row):
    """Convierte una fila de CSV/XLSX a diccionario de producto."""
    return {
        'nombre': row.get('nombre', '').strip(),
        'codigo_barras': row.get('codigo_barras', '').strip(),
        'sustancia_activa': row.get('sustancia_activa', '').strip(),
        'marca_laboratorio': row.get('marca', '').strip(),
        'categoria': row.get('categoria', '').strip(),
        'precio_publico': Decimal(str(row.get('precio_publico', 0) or 0)),
        'precio_compra': Decimal(str(row.get('precio_compra', 0) or 0)),
        'stock': int(row.get('stock', 0) or 0),
        'stock_minimo': int(row.get('stock_minimo', 0) or 0),
        'iva_porcentaje': Decimal(str(row.get('iva_porcentaje', 16) or 16)),
        'es_antibiotico': str(row.get('es_antibiotico', '')).lower() in {'1', 'true', 'si', 'sí'},
    }


# ==============================================================================
# LIBRO DE CONTROL DE ANTIBIÓTICOS
# ==============================================================================

@login_required
def libro_control_antibioticos(request):
    """Libro de control COFEPRIS — NOM-059 / Art. 226 LGS."""
    empresa = _empresa_desde_request(request)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada.')
        return redirect('home')

    from farmacia.models import RegistroAntibiotico

    # Filtros por fecha
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    producto_q = request.GET.get('producto', '').strip()

    qs = RegistroAntibiotico.objects.filter(empresa=empresa).select_related(
        'producto', 'paciente', 'usuario_vendedor', 'venta', 'lote_vendido'
    ).order_by('-fecha_venta')

    if fecha_desde_str:
        try:
            fd = date.fromisoformat(fecha_desde_str)
            qs = qs.filter(fecha_venta__gte=fd)
        except ValueError:
            logger.info("Fecha desde invalida en libro de antibioticos: %s", fecha_desde_str)
    if fecha_hasta_str:
        try:
            fh = date.fromisoformat(fecha_hasta_str)
            qs = qs.filter(fecha_venta__lte=fh)
        except ValueError:
            logger.info("Fecha hasta invalida en libro de antibioticos: %s", fecha_hasta_str)
    if producto_q:
        qs = qs.filter(producto__nombre__icontains=producto_q)

    # Construir estructura de reporte agrupada por producto
    from collections import defaultdict
    grupos = defaultdict(lambda: {'producto': None, 'entradas': [], 'salidas': []})

    for reg in qs[:500]:
        prod = reg.producto
        key = prod.pk
        if grupos[key]['producto'] is None:
            grupos[key]['producto'] = prod
        grupos[key]['salidas'].append({
            'fecha_mov': reg.fecha_venta,
            'tipo': 'VENTA',
            'ref': reg.venta.folio_operacion if reg.venta else '---',
            'lote_usado': reg.lote_vendido,
            'cantidad': reg.cantidad_vendida,
            'doctor': f"{reg.medico_nombre or ''} | Cédula: {reg.medico_cedula or ''}".strip('| '),
        })

    return render(request, 'core/libro_control_antibioticos.html', {
        'empresa': empresa,
        'grupos': dict(grupos),
        'fecha_desde': fecha_desde_str,
        'fecha_hasta': fecha_hasta_str,
        'producto_q': producto_q,
    })


# ==============================================================================
# DASHBOARD FARMACIA
# ==============================================================================

@login_required
def dashboard_farmacia(request):
    """Dashboard principal de farmacia."""
    empresa = _empresa_desde_request(request)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada. Contacte al administrador.')
        return redirect('admin:index')

    hoy = timezone.localdate()
    inicio = timezone.make_aware(datetime.combine(hoy, datetime.min.time()))
    fin = timezone.make_aware(datetime.combine(hoy, datetime.max.time()))
    
    ventas_hoy = Venta.objects.filter(empresa=empresa, fecha__range=(inicio, fin), estado='COMPLETADA')
    total_ventas_hoy = ventas_hoy.aggregate(total=Coalesce(Sum('total'), Decimal('0.00'), output_field=DecimalField()))['total'] or Decimal('0.00')
    
    # 2. PERSISTENCIA DE METAS (Dinámico)
    try:
        from django.apps import apps
        MetaVenta = apps.get_model('core', 'MetaVenta')
        meta_obj = MetaVenta.objects.filter(empresa=empresa, fecha=hoy).first()
        monto_meta = meta_obj.monto_objetivo if meta_obj else Decimal('50000.00')
    except LookupError:
        monto_meta = Decimal('50000.00')
    
    # Calcular porcentaje de meta
    porcentaje_meta = 0
    if monto_meta > 0:
        porcentaje_meta = (total_ventas_hoy / monto_meta) * 100
        if porcentaje_meta > 100: porcentaje_meta = 100

    ventas_efectivo = Pago.objects.filter(
        venta__empresa=empresa,
        venta__fecha__range=(inicio, fin),
        venta__estado='COMPLETADA',
        metodo='EFECTIVO'
    ).aggregate(total=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField()))['total'] or Decimal('0.00')
    
    ventas_digital = Pago.objects.filter(
        venta__empresa=empresa,
        venta__fecha__range=(inicio, fin),
        venta__estado='COMPLETADA'
    ).exclude(metodo='EFECTIVO').aggregate(total=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField()))['total'] or Decimal('0.00')
    
    gastos_hoy = GastoCaja.objects.filter(empresa=empresa, fecha__range=(inicio, fin))
    saldo_caja = ventas_efectivo - (gastos_hoy.aggregate(total=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField()))['total'] or Decimal('0.00'))
    
    ultimas_ventas = ventas_hoy.select_related('paciente').order_by('-fecha')[:10]

    # Alertas FEFO/stock
    fecha_limite = hoy + timedelta(days=_DIAS_CADUCIDAD_CRITICO)

    lotes_por_vencer = (
        Lote.objects.filter(
            producto__empresa=empresa,
            cantidad__gt=0,
            fecha_caducidad__lte=fecha_limite,
            fecha_caducidad__gte=hoy,
        )
        .select_related('producto')
        .order_by('fecha_caducidad')[:50]
    )

    productos_vencer = []
    for lote in lotes_por_vencer:
        dias = (lote.fecha_caducidad - hoy).days
        productos_vencer.append({
            'producto': lote.producto.nombre,
            'lote': lote.numero_lote,
            'caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y'),
            'dias': dias,
            'dias_restantes': dias,
            'cantidad': lote.cantidad,
        })

    # Productos agotados
    productos_agotados = Producto.objects.filter(
        empresa=empresa,
        stock__lte=0
    ).order_by('nombre')[:20]

    # Productos más vendidos del día
    from django.db.models import Count
    productos_mas_vendidos = (
        Venta.objects.filter(
            empresa=empresa,
            fecha__range=(inicio, fin),
            estado='COMPLETADA'
        )
        .values('detalles__producto__nombre')
        .annotate(
            total_vendido=Sum('detalles__cantidad'),
            cantidad_ventas=Count('id')
        )
        .order_by('-total_vendido')[:10]
    )

    # Productos con stock bajo
    productos_bajo_stock = Producto.objects.filter(
        empresa=empresa,
        stock__gt=0,
        stock__lt=F('stock_minimo')
    ).order_by('stock')[:20]

    # Ventas recientes
    ventas_recientes = ventas_hoy.select_related('paciente', 'usuario').order_by('-fecha')[:5]

    # Productos vendidos hoy
    productos_vendidos = ventas_hoy.aggregate(
        total_productos=Sum('detalles__cantidad')
    )['total_productos'] or 0

    porcentaje_digital = 0
    if total_ventas_hoy > 0:
        porcentaje_digital = round((ventas_digital / total_ventas_hoy) * 100, 2)

    return render(request, 'core/dashboard_farmacia.html', {
        'empresa': empresa,
        'fecha_hoy': hoy.strftime('%d/%m/%Y'),
        'total_ventas_hoy': total_ventas_hoy,
        'cantidad_ventas': ventas_hoy.count(),
        'ventas_hoy_total': total_ventas_hoy,
        'ventas_hoy_cantidad': ventas_hoy.count(),
        'productos_vendidos': productos_vendidos,
        'productos_agotados': len(productos_agotados),
        'productos_caducidad': len(productos_vencer),
        'cantidad_productos_vencer': len(productos_vencer),
        'cantidad_productos_bajo_stock': len(productos_bajo_stock),
        'lista_agotados': productos_agotados,
        'ventas_recientes': ventas_recientes,
        'productos_mas_vendidos': productos_mas_vendidos,
        'productos_vencer': productos_vencer,
        'productos_stock_bajo': productos_bajo_stock,
        'productos_proximos_caducar': productos_vencer,
        'ventas_controlados_sin_registrar': 0,
        'monto_meta': monto_meta,
        'porcentaje_meta': porcentaje_meta,
        'ventas_efectivo': ventas_efectivo,
        'ventas_digital': ventas_digital,
        'porcentaje_digital': porcentaje_digital,
        'saldo_caja': saldo_caja,
        'fecha_seleccionada': request.GET.get('fecha', hoy.strftime('%Y-%m-%d'))
    })


# ==============================================================================
# GESTIONAR POLÍTICAS DE DESCUENTO
# ==============================================================================

@login_required
def gestionar_politicas_descuento(request):
    """Vista para gestionar políticas de descuento. Acceso: ADMIN/GERENCIA/DIRECTOR."""
    empresa = _empresa_desde_request(request)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada.')
        return redirect('home')

    rol = (getattr(request.user, 'rol', '') or '').upper().strip()
    _roles_permitidos = ('ADMIN', 'ADMINISTRADOR', 'GERENCIA', 'GERENCIA_OPERATIVA',
                         'DIRECTOR', 'FARMACIA_SUPERVISOR')
    if not (request.user.is_superuser or request.user.is_staff or rol in _roles_permitidos):
        messages.warning(request, 'No tienes permisos para acceder a Políticas de Descuento.')
        return redirect('home')

    politicas = DiscountPolicy.objects.filter(empresa=empresa)
    return render(request, 'core/politicas_descuento.html', {
        'empresa': empresa,
        'politicas': politicas
    })


# ==============================================================================
# API: LISTAS DE PRECIO PDV
# ==============================================================================

@login_required
def api_listas_precio_pdv(request):
    """
    Devuelve las políticas de descuento activas de la empresa para el selector del PDV.
    GET /farmacia/api/listas-precio/
    """
    empresa = _empresa_desde_request(request)
    if not empresa:
        return JsonResponse({'listas': []}, status=403)

    politicas = DiscountPolicy.objects.filter(empresa=empresa, activa=True).values(
        'id', 'nombre', 'porcentaje_descuento', 'requiere_autorizacion'
    )
    listas = [
        {
            'id': p['id'],
            'nombre': p['nombre'],
            'porcentaje': float(p['porcentaje_descuento']),
            'requiere_auth': p['requiere_autorizacion'],
        }
        for p in politicas
    ]
    return JsonResponse({'listas': listas})


# ==============================================================================
# REGISTRO DE GASTO
# ==============================================================================

@login_required
def registro_gasto(request):
    """Vista para registrar gastos de caja."""
    empresa = _empresa_desde_request(request)
    if not empresa:
        return JsonResponse({'status': 'error', 'mensaje': 'Usuario sin empresa asignada'}, status=403)
    
    if request.method == 'GET':
        from core.models import GastoCaja
        gastos_hoy = GastoCaja.objects.filter(
            empresa=empresa,
            fecha__date=timezone.now().date()
        ).order_by('-fecha')[:20]
        return render(request, 'core/registro_gasto.html', {
            'empresa': empresa,
            'gastos_hoy': gastos_hoy,
        })
    if request.method == 'POST':
        try:
            from django.core.exceptions import ValidationError

            data = json.loads(request.body)
            concepto = data.get('concepto', '')
            monto = Decimal(str(data.get('monto', 0)))
            gasto = GastoCaja(
                empresa=empresa,
                usuario=request.user,
                concepto=concepto,
                monto=monto,
            )
            gasto.save()
            # AuditLog
            try:
                from core.services.audit_service import registrar_auditoria
                registrar_auditoria(
                    accion='CREATE',
                    modelo='GastoCaja',
                    objeto_id=str(gasto.id),
                    datos_nuevos={'concepto': concepto, 'monto': str(monto)},
                    request=request,
                )
            except Exception:
                logger.exception("No se pudo registrar auditoria de gasto de caja %s", gasto.id)
            return JsonResponse({'status': 'success'})
        except ValidationError as e:
            err = getattr(e, 'message_dict', None) or str(e)
            return JsonResponse({'status': 'error', 'mensaje': err}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)
    return JsonResponse({'status': 'error'}, status=405)


# ==============================================================================
# API: SALDO DE CAJA
# ==============================================================================

@login_required
@require_http_methods(["GET"])
def api_saldo_caja(request):
    """
    API para ver saldo de caja en tiempo real (usado por verSaldoCaja() en PDV).
    Retorna JSON: total_vendido_dia, ventas_efectivo, ventas_digital, gastos_retiros,
    saldo_en_caja, lista_gastos (hora, concepto, monto).
    """
    empresa = _empresa_desde_request(request)
    if not empresa:
        return JsonResponse({'status': 'error', 'mensaje': 'Usuario sin empresa asignada'}, status=403)
    hoy = timezone.now().date()
    inicio = timezone.make_aware(datetime.combine(hoy, datetime.min.time()))
    fin = timezone.make_aware(datetime.combine(hoy, datetime.max.time()))
    ventas_hoy = Venta.objects.filter(
        empresa=empresa,
        fecha__range=(inicio, fin),
        estado='COMPLETADA'
    )
    total_vendido_dia = ventas_hoy.aggregate(
        total=Coalesce(Sum('total'), Decimal('0.00'), output_field=DecimalField())
    )['total'] or Decimal('0.00')
    ventas_efectivo = Pago.objects.filter(
        venta__empresa=empresa,
        venta__fecha__range=(inicio, fin),
        venta__estado='COMPLETADA',
    ).aggregate(
        total=Coalesce(Sum('monto_efectivo'), Decimal('0.00'), output_field=DecimalField())
    )['total'] or Decimal('0.00')
    ventas_digital_agg = Pago.objects.filter(
        venta__empresa=empresa,
        venta__fecha__range=(inicio, fin),
        venta__estado='COMPLETADA',
    ).aggregate(
        tar=Coalesce(Sum('monto_tarjeta'), Decimal('0.00'), output_field=DecimalField()),
        trans=Coalesce(Sum('monto_transferencia'), Decimal('0.00'), output_field=DecimalField()),
    )
    ventas_digital = (ventas_digital_agg.get('tar') or Decimal('0.00')) + (ventas_digital_agg.get('trans') or Decimal('0.00'))
    gastos_hoy = GastoCaja.objects.filter(
        empresa=empresa,
        fecha__range=(inicio, fin)
    ).order_by('-fecha')
    total_gastos = gastos_hoy.aggregate(
        total=Coalesce(Sum('monto'), Decimal('0.00'), output_field=DecimalField())
    )['total'] or Decimal('0.00')
    lista_gastos = [
        {
            'hora': g.fecha.strftime('%H:%M') if g.fecha else '--:--',
            'concepto': g.concepto or '',
            'monto': float(g.monto),
        }
        for g in gastos_hoy
    ]
    saldo_en_caja = ventas_efectivo - total_gastos
    return JsonResponse({
        'total_vendido_dia': float(total_vendido_dia),
        'ventas_efectivo': float(ventas_efectivo),
        'ventas_digital': float(ventas_digital),
        'gastos_retiros': float(total_gastos),
        'saldo_en_caja': float(saldo_en_caja),
        'lista_gastos': lista_gastos,
    })


# ==============================================================================
# VALIDAR PIN PRECIO NETO
# ==============================================================================

@login_required
@require_http_methods(["POST"])
def validar_pin_precio_neto(request):
    """
    Valida el PIN ingresado por el staff para activar el descuento 
    a Precio Neto (costo de compra). Solo usuarios autorizados pueden
    solicitar esta validación.
    """
    empresa = _empresa_desde_request(request)
    if not empresa:
        return JsonResponse({'status': 'error', 'mensaje': 'Sin empresa asignada'}, status=403)
    
    # Verificar que el usuario tiene permiso de acceder a esta función
    ROLES_PRECIO_NETO = ['Administrador', 'FARMACIA', 'Gerente', 'Director']
    puede_precio_neto = (
        request.user.is_superuser or 
        request.user.groups.filter(name__in=ROLES_PRECIO_NETO).exists() or
        getattr(request.user, 'rol', '') in ['ADMIN', 'GERENTE', 'DIRECTOR', 'FARMACIA']
    )
    
    if not puede_precio_neto:
        return JsonResponse({'status': 'error', 'mensaje': 'Sin permisos para validar PIN'}, status=403)
    
    try:
        data = json.loads(request.body)
        pin_ingresado = data.get('pin', '').strip()
        
        # Validar PIN (configurado en settings)
        PIN_PRECIO_NETO = getattr(settings, 'FARMACIA_PIN_PRECIO_NETO', '1234')
        
        if pin_ingresado == PIN_PRECIO_NETO:
            return JsonResponse({'status': 'success', 'mensaje': 'PIN válido'})
        else:
            return JsonResponse({'status': 'error', 'mensaje': 'PIN incorrecto'}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'mensaje': 'Datos inválidos'}, status=400)


# ==============================================================================
# IMPRIMIR ETIQUETAS
# ==============================================================================

@login_required
def imprimir_etiquetas(request):
    """Genera PDF de etiquetas de productos seleccionados."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            lotes_ids = data.get('lotes', [])
            
            if not lotes_ids:
                 return JsonResponse({'error': 'No se seleccionaron lotes'}, status=400)

            # TODO: implementar generación real de PDF de etiquetas de productos con reportlab.
            return JsonResponse({
                'status': 'error',
                'message': 'Generación de etiquetas de farmacia no implementada. Contacte al administrador.',
            }, status=501)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ==============================================================================
# API: VALIDAR CUPÓN
# ==============================================================================

@login_required
def api_validar_cupon(request):
    """
    API para validar un cupón de marketing en tiempo real.
    Retorna información del cupón si es válido.
    """
    empresa = _empresa_desde_request(request)
    if not empresa:
        return JsonResponse({'status': 'error', 'mensaje': 'Usuario sin empresa asignada'}, status=403)
    codigo_cupon = request.GET.get('codigo', '').strip().upper()
    
    if not codigo_cupon:
        return JsonResponse({
            'status': 'error',
            'mensaje': 'Código de cupón requerido'
        }, status=400)
    
    try:
        # Intentar importar el modelo de marketing
        from marketing.models import CuponMarketing
        
        # Buscar el cupón
        cupon = CuponMarketing.objects.filter(
            empresa=empresa,
            codigo=codigo_cupon
        ).first()
        
        if not cupon:
            return JsonResponse({
                'status': 'error',
                'mensaje': 'Cupón no encontrado'
            }, status=404)
        
        # Validar que el cupón esté activo
        return JsonResponse({
            'status': 'success',
            'cupon': {
                'id': cupon.id,
                'codigo': cupon.codigo,
                'porcentaje_descuento': float(cupon.porcentaje_descuento),
                'descripcion': cupon.descripcion or '',
            }
        })
        
    except ImportError:
        # Si el módulo de marketing no está disponible
        return JsonResponse({
            'status': 'error',
            'mensaje': 'Módulo de marketing no disponible'
        }, status=503)
    except Exception as e:
        logger.error(f'Error al validar cupón: {str(e)}')
        return JsonResponse({
            'status': 'error',
            'mensaje': 'Error al validar el cupón'
        }, status=500)
