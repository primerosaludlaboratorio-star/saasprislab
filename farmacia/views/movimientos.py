"""
Vistas de Gestión de Movimientos de Inventario y Kardex para Farmacia
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, ListView
from django.conf import settings
from django.db.models import Sum, F, Q, DecimalField, Value
from django.db.models.functions import Coalesce
from django.db import DatabaseError
from django.core.exceptions import ValidationError
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.contrib import messages
from datetime import timedelta, date
from decimal import Decimal

from core.models import Producto, Lote
from farmacia.models import MovimientoInventario, MotivoAjuste
from core.utils.sucursal_helpers import get_user_primary_sucursal

# Umbrales de caducidad configurables vía settings
_DIAS_CADUCIDAD_CRITICO = getattr(settings, 'FARMACIA_DIAS_CADUCIDAD_CRITICO', 30)
_DIAS_CADUCIDAD_ALERTA = getattr(settings, 'FARMACIA_DIAS_CADUCIDAD_ALERTA', 90)


class FarmaciaAlertasView(LoginRequiredMixin, TemplateView):
    """
    Dashboard de Alertas Proactivas - Semáforo de Gestión.
    """
    template_name = 'farmacia/dashboard_alertas.html'
    
    def get(self, request, *args, **kwargs):
        if not getattr(request.user, 'empresa', None):
            messages.error(request, 'Usuario no tiene empresa asignada. Contacte al administrador.')
            return redirect('home')
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa = getattr(self.request.user, 'empresa', None)
        if not empresa:
            return context
        hoy = date.today()
        
        # PANEL 1: SEMÁFORO DE CADUCIDAD
        lotes_todos = Lote.objects.filter(
            producto__empresa=empresa,
            cantidad__gt=0
        ).select_related('producto').order_by('fecha_caducidad')
        
        fecha_critico = hoy + timedelta(days=_DIAS_CADUCIDAD_CRITICO)
        lotes_criticos = lotes_todos.filter(
            fecha_caducidad__lte=fecha_critico
        )

        fecha_alerta_inicio = hoy + timedelta(days=_DIAS_CADUCIDAD_CRITICO + 1)
        fecha_alerta_fin = hoy + timedelta(days=_DIAS_CADUCIDAD_ALERTA)
        lotes_alerta = lotes_todos.filter(
            fecha_caducidad__gt=fecha_alerta_inicio,
            fecha_caducidad__lte=fecha_alerta_fin
        )
        
        context['lotes_criticos'] = lotes_criticos
        context['lotes_alerta'] = lotes_alerta
        context['total_criticos'] = lotes_criticos.count()
        context['total_alerta'] = lotes_alerta.count()
        
        valor_riesgo_critico = sum(
            (l.cantidad * l.costo_adquisicion) for l in lotes_criticos
        )
        context['valor_riesgo_critico'] = valor_riesgo_critico
        
        # PANEL 2: STOCK BAJO
        productos_stock_bajo = Producto.objects.filter(
            empresa=empresa,
            stock__gt=0
        ).annotate(
            stock_minimo_calc=F('stock_minimo')
        ).filter(
            stock__lt=F('stock_minimo')
        ).order_by('stock')[:20]
        
        context['productos_stock_bajo'] = productos_stock_bajo
        context['total_stock_bajo'] = productos_stock_bajo.count()
        
        # PANEL 3: PRODUCTOS YA CADUCADOS
        lotes_caducados = Lote.objects.filter(
            producto__empresa=empresa,
            cantidad__gt=0,
            fecha_caducidad__lt=hoy
        ).select_related('producto').order_by('fecha_caducidad')
        
        context['lotes_caducados'] = lotes_caducados
        context['total_caducados'] = lotes_caducados.count()
        
        valor_perdido = sum(
            (l.cantidad * l.costo_adquisicion) for l in lotes_caducados
        )
        context['valor_perdido'] = valor_perdido
        
        # PANEL 4: DEMANDA INSATISFECHA
        try:
            from core.models import DemandaInsatisfecha
            fecha_inicio_demanda = timezone.now() - timedelta(days=30)
            
            demandas = DemandaInsatisfecha.objects.filter(
                empresa=empresa,
                fecha__gte=fecha_inicio_demanda
            ).values('producto_nombre', 'causa').annotate(
                total_cantidad=Sum('cantidad_dejada')
            ).order_by('-total_cantidad')[:15]
            
            context['demandas_insatisfechas'] = demandas
            context['total_demandas'] = demandas.count()
        except ImportError:
            context['demandas_insatisfechas'] = []
            context['total_demandas'] = 0
        
        context['total_alertas'] = (
            context['total_criticos'] +
            context['total_alerta'] +
            context['total_stock_bajo'] +
            context['total_caducados']
        )
        context['fecha_hoy'] = hoy
        return context


class KardexListView(LoginRequiredMixin, ListView):
    """
    Vista de lista de movimientos del Kardex con filtros.
    """
    model = MovimientoInventario
    template_name = 'farmacia/kardex_list.html'
    context_object_name = 'movimientos'
    paginate_by = 50
    
    def get_queryset(self):
        empresa = getattr(self.request.user, 'empresa', None)
        if not empresa:
            return MovimientoInventario.objects.none()
        queryset = MovimientoInventario.objects.filter(
            empresa=empresa
        ).select_related(
            'producto', 'lote', 'usuario_responsable', 'proveedor', 'venta'
        ).order_by('-fecha_movimiento')

        producto_q = self.request.GET.get('producto_q', '').strip()
        if producto_q:
            queryset = queryset.filter(
                Q(producto__nombre__icontains=producto_q) |
                Q(producto__codigo_barras__icontains=producto_q) |
                Q(folio__icontains=producto_q)
            )

        tipo = self.request.GET.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo_movimiento=tipo)

        fecha_desde = self.request.GET.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fecha_movimiento__gte=fecha_desde)

        fecha_hasta = self.request.GET.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fecha_movimiento__lte=fecha_hasta)

        return queryset

    def get(self, request, *args, **kwargs):
        if request.GET.get('formato') == 'excel':
            return self._exportar_excel(request)
        return super().get(request, *args, **kwargs)

    def _exportar_excel(self, request):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        qs = self.get_queryset()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Kardex'
        headers = ['Folio', 'Fecha', 'Tipo', 'Producto', 'Lote', 'Cantidad',
                   'Costo Unitario', 'Stock Anterior', 'Stock Resultante', 'Usuario', 'Observaciones']
        ws.append(headers)
        fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.fill = fill
            cell.font = Font(color='FFFFFF', bold=True)

        for mov in qs[:5000]:
            ws.append([
                mov.folio,
                mov.fecha_movimiento.strftime('%d/%m/%Y %H:%M') if mov.fecha_movimiento else '',
                mov.get_tipo_movimiento_display(),
                mov.producto.nombre if mov.producto else '',
                mov.lote.numero_lote if mov.lote else '',
                float(mov.cantidad),
                float(mov.costo_unitario or 0),
                float(mov.stock_anterior or 0),
                float(mov.stock_resultante or 0),
                mov.usuario_responsable.get_full_name() if mov.usuario_responsable else '',
                mov.observaciones or '',
            ])

        for col_letter, width in [('A', 15), ('B', 16), ('C', 20), ('D', 35), ('E', 18),
                                    ('F', 10), ('G', 14), ('H', 14), ('I', 14), ('J', 22), ('K', 35)]:
            ws.column_dimensions[col_letter].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="kardex_movimientos.xlsx"'
        return resp

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipos_movimiento'] = MovimientoInventario.TIPO_MOVIMIENTO
        return context


@login_required
@permission_required('farmacia.add_movimientoinventario', raise_exception=True)
def crear_movimiento_manual(request):
    """
    Vista para crear movimientos manuales de inventario.
    """
    if request.method == 'POST':
        try:
            producto_id = request.POST.get('producto_id')
            lote_id = request.POST.get('lote_id')
            tipo_movimiento = request.POST.get('tipo_movimiento')
            cantidad = Decimal(request.POST.get('cantidad', 0))
            costo_unitario = Decimal(request.POST.get('costo_unitario', 0))
            motivo_ajuste_id = request.POST.get('motivo_ajuste_id')
            observaciones = request.POST.get('observaciones', '')
            
            if not producto_id or cantidad <= 0:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': 'Datos incompletos o inválidos'
                }, status=400)
            
            empresa = getattr(request.user, 'empresa', None)
            producto = get_object_or_404(Producto, id=producto_id, empresa=empresa)
            lote = get_object_or_404(Lote, id=lote_id, producto__empresa=empresa) if lote_id else None
            motivo_ajuste = get_object_or_404(MotivoAjuste, id=motivo_ajuste_id) if motivo_ajuste_id else None
            sucursal = get_user_primary_sucursal(request.user)
            
            movimiento = MovimientoInventario(
                empresa=empresa,
                sucursal=sucursal,
                producto=producto,
                lote=lote,
                tipo_movimiento=tipo_movimiento,
                cantidad=cantidad,
                costo_unitario=costo_unitario,
                usuario_responsable=request.user,
                motivo_ajuste=motivo_ajuste,
                observaciones=observaciones
            )
            
            if motivo_ajuste and motivo_ajuste.requiere_autorizacion_gerente:
                movimiento.requiere_autorizacion = True
            
            movimiento.save()
            
            return JsonResponse({
                'status': 'success',
                'mensaje': f'Movimiento {movimiento.folio} creado exitosamente',
                'folio': movimiento.folio,
                'stock_nuevo': float(movimiento.stock_resultante)
            })
            
        except (DatabaseError, ValueError, TypeError, ValidationError) as e:
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error al crear movimiento: {str(e)}'
            }, status=500)
    
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada.')
        return redirect('dashboard')
    productos = Producto.objects.filter(empresa=empresa).order_by('nombre')[:100]
    motivos = MotivoAjuste.objects.filter(empresa=empresa, activo=True)
    
    return render(request, 'farmacia/crear_movimiento.html', {
        'productos': productos,
        'motivos': motivos,
        'tipos_movimiento': MovimientoInventario.TIPO_MOVIMIENTO
    })


@login_required
@permission_required('farmacia.autorizar_movimientos', raise_exception=True)
def autorizar_movimiento(request, movimiento_id):
    """
    Autorizar un movimiento de inventario que requiere validación.
    """
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        return JsonResponse({'status': 'error', 'mensaje': 'Usuario sin empresa asignada'}, status=403)
    movimiento = get_object_or_404(
        MovimientoInventario,
        id=movimiento_id,
        empresa=empresa
    )
    
    if not movimiento.requiere_autorizacion:
        return JsonResponse({
            'status': 'error',
            'mensaje': 'Este movimiento no requiere autorización'
        }, status=400)
    
    if movimiento.autorizado:
        return JsonResponse({
            'status': 'error',
            'mensaje': 'Este movimiento ya fue autorizado'
        }, status=400)
    
    MovimientoInventario.objects.filter(pk=movimiento.pk).update(
        autorizado=True,
        autorizado_por=request.user,
        fecha_autorizacion=timezone.now()
    )
    
    return JsonResponse({
        'status': 'success',
        'mensaje': f'Movimiento {movimiento.folio} autorizado exitosamente'
    })


@login_required
def api_lotes_producto(request, producto_id):
    """
    API para obtener lotes disponibles de un producto.
    """
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        return JsonResponse({'lotes': [], 'producto': None})

    try:
        producto = Producto.objects.get(id=producto_id, empresa=empresa)
    except Producto.DoesNotExist:
        return JsonResponse({'lotes': [], 'producto': None}, status=404)

    today = date.today()
    lotes = Lote.objects.filter(
        producto=producto,
        cantidad__gt=0
    ).order_by('fecha_caducidad')

    lotes_data = []
    lote_proximo = None
    for l in lotes:
        dias = (l.fecha_caducidad - today).days if l.fecha_caducidad else None
        es_vencido = (l.fecha_caducidad is not None and l.fecha_caducidad < today)
        costo = float(l.costo_adquisicion) if l.costo_adquisicion else 0
        lotes_data.append({
            'id': l.id,
            'numero_lote': l.numero_lote,
            'fecha_caducidad': l.fecha_caducidad.strftime('%Y-%m-%d') if l.fecha_caducidad else None,
            'cantidad': float(l.cantidad),
            'costo_adquisicion': costo,
            'dias_restantes': dias,
            'es_vencido': es_vencido,
        })
        if lote_proximo is None and not es_vencido:
            lote_proximo = l

    stock_vigente = sum(float(l['cantidad']) for l in lotes_data if not l.get('es_vencido'))
    stock_total = sum(float(l['cantidad']) for l in lotes_data)
    lotes_vencidos_count = sum(1 for l in lotes_data if l.get('es_vencido'))

    precio_publico = float(producto.precio_publico)
    costo_lote_proximo = float(lote_proximo.costo_adquisicion) if (lote_proximo and lote_proximo.costo_adquisicion) else 0
    alerta_precio_bajo = (costo_lote_proximo > 0 and precio_publico < costo_lote_proximo)

    producto_data = {
        'id': producto.id,
        'nombre_comercial': producto.nombre,
        'sustancia_activa': producto.sustancia_activa or '',
        'codigo_barras': producto.codigo_barras or '',
        'precio_base': precio_publico,
        'precio_venta': precio_publico,
        'precio_compra': float(producto.precio_compra) if producto.precio_compra else 0,
        'iva_pct': float(producto.iva_porcentaje) if producto.iva_porcentaje else 0,
        'es_antibiotico': producto.es_antibiotico if hasattr(producto, 'es_antibiotico') else False,
        'es_controlado': producto.es_antibiotico if hasattr(producto, 'es_antibiotico') else False,
        'stock': int(producto.stock) if producto.stock else 0,
        'stock_total': stock_vigente,
        'stock_total_fisico': stock_total,
        'lotes_vencidos_count': lotes_vencidos_count,
        'alerta_precio_bajo': alerta_precio_bajo,
        'costo_lote': costo_lote_proximo,
        'lote_id': lote_proximo.id if lote_proximo else None,
        'numero_lote_proximo': lote_proximo.numero_lote if lote_proximo else '',
        'proxima_caducidad': lote_proximo.fecha_caducidad.strftime('%Y-%m-%d') if lote_proximo and lote_proximo.fecha_caducidad else None,
        'sin_stock_vigente': lote_proximo is None,
    }

    return JsonResponse({'lotes': lotes_data, 'producto': producto_data})


@login_required
def reporte_valorizacion_inventario(request):
    """
    Reporte financiero de valorización del inventario.
    """
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada.')
        return redirect('dashboard')
    
    productos = Producto.objects.filter(
        empresa=empresa,
        stock__gt=0
    ).annotate(
        valor_inventario=F('stock') * F('precio_compra')
    ).order_by('-valor_inventario')
    
    total_unidades = productos.aggregate(
        total=Coalesce(Sum('stock'), Value(0), output_field=DecimalField())
    )['total']
    
    total_valor = productos.aggregate(
        total=Coalesce(Sum('valor_inventario'), Value(0), output_field=DecimalField())
    )['total']
    
    if request.GET.get('formato') == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Valorización'
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Reporte de Valorización — {empresa.nombre}'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f'Fecha: {timezone.now().strftime("%d/%m/%Y %H:%M")}'

        ws.append([])
        headers = ['Producto', 'SKU', 'Categoría', 'Stock', 'Costo Unitario', 'Valor Total']
        ws.append(headers)
        fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        for col in range(1, 7):
            cell = ws.cell(ws.max_row, col)
            cell.fill = fill
            cell.font = Font(color='FFFFFF', bold=True)

        for p in productos:
            cat_val = p.categoria if isinstance(p.categoria, str) else (
                p.categoria.nombre if hasattr(p.categoria, 'nombre') else str(p.categoria or '')
            )
            ws.append([
                p.nombre,
                p.codigo_barras or '',
                cat_val,
                float(p.stock),
                float(p.precio_compra or 0),
                float(p.valor_inventario or 0),
            ])

        ws.append([])
        ws.append(['', '', '', float(total_unidades), '', float(total_valor)])
        ws.cell(ws.max_row, 4).font = Font(bold=True)
        ws.cell(ws.max_row, 6).font = Font(bold=True)

        for col, width in [('A', 35), ('B', 15), ('C', 20), ('D', 10), ('E', 18), ('F', 18)]:
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="valorizacion_inventario.xlsx"'
        return resp

    return render(request, 'farmacia/reporte_valorizacion.html', {
        'productos': productos,
        'total_unidades': total_unidades,
        'total_valor': total_valor,
        'fecha_reporte': timezone.now()
    })
