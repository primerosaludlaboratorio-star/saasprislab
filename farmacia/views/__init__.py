"""
Vistas de Gestión Avanzada para Farmacia
Dashboard de Alertas Proactivas y Gestión de Kardex
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import TemplateView, ListView, CreateView, FormView
from django.conf import settings
from django.db.models import Sum, F, Q, Count, DecimalField, Value
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db import transaction
from django.contrib import messages
from datetime import timedelta, date, datetime, time
from decimal import Decimal
import json

# Umbrales de caducidad configurables vía settings
_DIAS_CADUCIDAD_CRITICO = getattr(settings, 'FARMACIA_DIAS_CADUCIDAD_CRITICO', 30)
_DIAS_CADUCIDAD_ALERTA = getattr(settings, 'FARMACIA_DIAS_CADUCIDAD_ALERTA', 90)

from core.models import Producto, Lote, Venta, Pago
from farmacia.models import MovimientoInventario, Proveedor, MotivoAjuste
from farmacia.forms import (
    RegistrarCompraForm, DetalleCompraForm, CorteCajaFarmaciaForm,
    AjusteInventarioForm, GenerarEtiquetasForm
)


# ==============================================================================
# DASHBOARD DE ALERTAS PROACTIVAS
# ==============================================================================
class FarmaciaAlertasView(LoginRequiredMixin, TemplateView):
    """
    Dashboard de Alertas Proactivas - Semáforo de Gestión.
    
    4 Paneles Críticos:
    1. Semáforo de Caducidad (Crítico 0-30, Alerta 31-90)
    2. Stock Bajo (por debajo del mínimo)
    3. Productos Caducados (YA VENCIDOS)
    4. Demanda Insatisfecha
    """
    template_name = 'farmacia/dashboard_alertas.html'
    
    def get(self, request, *args, **kwargs):
        if not getattr(request.user, 'empresa', None):
            messages.error(request, 'Usuario no tiene empresa asignada. Contacte al administrador.')
            return redirect('home')
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa = getattr(self.request.user, 'empresa', None)
        if not empresa:
            return context
        hoy = date.today()
        
        # ============================================================
        # PANEL 1: SEMÁFORO DE CADUCIDAD
        # ============================================================
        lotes_todos = Lote.objects.filter(
            producto__empresa=empresa,
            cantidad__gt=0
        ).select_related('producto').order_by('fecha_caducidad')
        
        # CRITICO (0-DIAS_CADUCIDAD_CRITICO días)
        fecha_critico = hoy + timedelta(days=_DIAS_CADUCIDAD_CRITICO)
        lotes_criticos = lotes_todos.filter(
            fecha_caducidad__lte=fecha_critico
        )

        # ALERTA (DIAS_CADUCIDAD_CRITICO+1 a DIAS_CADUCIDAD_ALERTA días)
        fecha_alerta_inicio = hoy + timedelta(days=_DIAS_CADUCIDAD_CRITICO + 1)
        fecha_alerta_fin = hoy + timedelta(days=_DIAS_CADUCIDAD_ALERTA)
        lotes_alerta = lotes_todos.filter(
            fecha_caducidad__gt=fecha_alerta_inicio,
            fecha_caducidad__lte=fecha_alerta_fin
        )
        
        context['lotes_criticos'] = lotes_criticos
        context['lotes_alerta'] = lotes_alerta
        context['total_criticos'] = lotes_criticos.count()
        context['total_alerta'] = lotes_alerta.count()
        
        # Valor en riesgo (costo total de lotes críticos)
        valor_riesgo_critico = sum(
            (l.cantidad * l.costo_adquisicion) for l in lotes_criticos
        )
        context['valor_riesgo_critico'] = valor_riesgo_critico
        
        # ============================================================
        # PANEL 2: STOCK BAJO (PUNTO DE REORDEN)
        # ============================================================
        # Productos con stock por debajo del mínimo
        productos_stock_bajo = Producto.objects.filter(
            empresa=empresa,
            stock__gt=0
        ).annotate(
            stock_minimo_calc=F('stock_minimo')
        ).filter(
            stock__lt=F('stock_minimo')
        ).order_by('stock')[:20]
        
        context['productos_stock_bajo'] = productos_stock_bajo
        context['total_stock_bajo'] = productos_stock_bajo.count()
        
        # ============================================================
        # PANEL 3: PRODUCTOS YA CADUCADOS
        # ============================================================
        lotes_caducados = Lote.objects.filter(
            producto__empresa=empresa,
            cantidad__gt=0,
            fecha_caducidad__lt=hoy
        ).select_related('producto').order_by('fecha_caducidad')
        
        context['lotes_caducados'] = lotes_caducados
        context['total_caducados'] = lotes_caducados.count()
        
        # Valor perdido por caducidad
        valor_perdido = sum(
            (l.cantidad * l.costo_adquisicion) for l in lotes_caducados
        )
        context['valor_perdido'] = valor_perdido
        
        # ============================================================
        # PANEL 4: DEMANDA INSATISFECHA (ÚLTIMOS 30 DÍAS)
        # ============================================================
        try:
            from core.models import DemandaInsatisfecha
            fecha_inicio_demanda = timezone.now() - timedelta(days=30)
            
            demandas = DemandaInsatisfecha.objects.filter(
                empresa=empresa,
                fecha__gte=fecha_inicio_demanda
            ).values('producto_nombre', 'causa').annotate(
                total_cantidad=Sum('cantidad_dejada')
            ).order_by('-total_cantidad')[:15]
            
            context['demandas_insatisfechas'] = demandas
            context['total_demandas'] = demandas.count()
        except ImportError:
            context['demandas_insatisfechas'] = []
            context['total_demandas'] = 0
        
        # ============================================================
        # RESUMEN GENERAL
        # ============================================================
        context['total_alertas'] = (
            context['total_criticos'] +
            context['total_alerta'] +
            context['total_stock_bajo'] +
            context['total_caducados']
        )
        
        context['fecha_hoy'] = hoy
        
        return context


# ==============================================================================
# GESTIÓN DE MOVIMIENTOS DE INVENTARIO (KARDEX)
# ==============================================================================
class KardexListView(LoginRequiredMixin, ListView):
    """
    Vista de lista de movimientos del Kardex con filtros.
    """
    model = MovimientoInventario
    template_name = 'farmacia/kardex_list.html'
    context_object_name = 'movimientos'
    paginate_by = 50
    
    def get_queryset(self):
        empresa = getattr(self.request.user, 'empresa', None)
        if not empresa:
            return MovimientoInventario.objects.none()
        queryset = MovimientoInventario.objects.filter(
            empresa=empresa
        ).select_related(
            'producto', 'lote', 'usuario_responsable', 'proveedor', 'venta'
        ).order_by('-fecha_movimiento')

        producto_q = self.request.GET.get('producto_q', '').strip()
        if producto_q:
            queryset = queryset.filter(
                Q(producto__nombre__icontains=producto_q) |
                Q(producto__codigo_barras__icontains=producto_q) |
                Q(folio__icontains=producto_q)
            )

        tipo = self.request.GET.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo_movimiento=tipo)

        fecha_desde = self.request.GET.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fecha_movimiento__gte=fecha_desde)

        fecha_hasta = self.request.GET.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fecha_movimiento__lte=fecha_hasta)

        return queryset

    def get(self, request, *args, **kwargs):
        # Exportación Excel
        if request.GET.get('formato') == 'excel':
            return self._exportar_excel(request)
        return super().get(request, *args, **kwargs)

    def _exportar_excel(self, request):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse

        qs = self.get_queryset()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Kardex'
        headers = ['Folio', 'Fecha', 'Tipo', 'Producto', 'Lote', 'Cantidad',
                   'Costo Unitario', 'Stock Anterior', 'Stock Resultante', 'Usuario', 'Observaciones']
        ws.append(headers)
        fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.fill = fill
            cell.font = Font(color='FFFFFF', bold=True)

        for mov in qs[:5000]:
            ws.append([
                mov.folio,
                mov.fecha_movimiento.strftime('%d/%m/%Y %H:%M') if mov.fecha_movimiento else '',
                mov.get_tipo_movimiento_display(),
                mov.producto.nombre if mov.producto else '',
                mov.lote.numero_lote if mov.lote else '',
                float(mov.cantidad),
                float(mov.costo_unitario or 0),
                float(mov.stock_anterior or 0),
                float(mov.stock_resultante or 0),
                mov.usuario_responsable.get_full_name() if mov.usuario_responsable else '',
                mov.observaciones or '',
            ])

        for col_letter, width in [('A', 15), ('B', 16), ('C', 20), ('D', 35), ('E', 18),
                                    ('F', 10), ('G', 14), ('H', 14), ('I', 14), ('J', 22), ('K', 35)]:
            ws.column_dimensions[col_letter].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="kardex_movimientos.xlsx"'
        return resp

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipos_movimiento'] = MovimientoInventario.TIPO_MOVIMIENTO
        return context


# ==============================================================================
# CREAR MOVIMIENTO MANUAL DE INVENTARIO (CON PERMISO)
# ==============================================================================
@login_required
@permission_required('farmacia.add_movimientoinventario', raise_exception=True)
def crear_movimiento_manual(request):
    """
    Vista para crear movimientos manuales de inventario.
    Requiere permiso farmacia.add_movimientoinventario.
    """
    if request.method == 'POST':
        try:
            producto_id = request.POST.get('producto_id')
            lote_id = request.POST.get('lote_id')
            tipo_movimiento = request.POST.get('tipo_movimiento')
            cantidad = Decimal(request.POST.get('cantidad', 0))
            costo_unitario = Decimal(request.POST.get('costo_unitario', 0))
            motivo_ajuste_id = request.POST.get('motivo_ajuste_id')
            observaciones = request.POST.get('observaciones', '')
            
            # Validaciones
            if not producto_id or cantidad <= 0:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': 'Datos incompletos o inválidos'
                }, status=400)
            
            empresa = getattr(request.user, 'empresa', None)
            producto = get_object_or_404(Producto, id=producto_id, empresa=empresa)
            lote = get_object_or_404(Lote, id=lote_id, producto__empresa=empresa) if lote_id else None
            motivo_ajuste = get_object_or_404(MotivoAjuste, id=motivo_ajuste_id) if motivo_ajuste_id else None
            
            # Crear movimiento
            movimiento = MovimientoInventario(
                empresa=empresa,
                sucursal=getattr(request.user, 'sucursal', None),
                producto=producto,
                lote=lote,
                tipo_movimiento=tipo_movimiento,
                cantidad=cantidad,
                costo_unitario=costo_unitario,
                usuario_responsable=request.user,
                motivo_ajuste=motivo_ajuste,
                observaciones=observaciones
            )
            
            # Validar si requiere autorización
            if motivo_ajuste and motivo_ajuste.requiere_autorizacion_gerente:
                movimiento.requiere_autorizacion = True
            
            # Guardar (el método save() maneja toda la lógica transaccional)
            movimiento.save()
            
            return JsonResponse({
                'status': 'success',
                'mensaje': f'Movimiento {movimiento.folio} creado exitosamente',
                'folio': movimiento.folio,
                'stock_nuevo': float(movimiento.stock_resultante)
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error al crear movimiento: {str(e)}'
            }, status=500)
    
    # GET: Mostrar formulario
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada.')
        return redirect('dashboard')
    productos = Producto.objects.filter(empresa=empresa).order_by('nombre')[:100]
    motivos = MotivoAjuste.objects.filter(empresa=empresa, activo=True)
    
    return render(request, 'farmacia/crear_movimiento.html', {
        'productos': productos,
        'motivos': motivos,
        'tipos_movimiento': MovimientoInventario.TIPO_MOVIMIENTO
    })


# ==============================================================================
# AUTORIZAR MOVIMIENTO (SOLO GERENTES)
# ==============================================================================
@login_required
@permission_required('farmacia.autorizar_movimientos', raise_exception=True)
def autorizar_movimiento(request, movimiento_id):
    """
    Autorizar un movimiento de inventario que requiere validación.
    Solo para usuarios con permiso farmacia.autorizar_movimientos.
    """
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        return JsonResponse({'status': 'error', 'mensaje': 'Usuario sin empresa asignada'}, status=403)
    movimiento = get_object_or_404(
        MovimientoInventario,
        id=movimiento_id,
        empresa=empresa
    )
    
    if not movimiento.requiere_autorizacion:
        return JsonResponse({
            'status': 'error',
            'mensaje': 'Este movimiento no requiere autorización'
        }, status=400)
    
    if movimiento.autorizado:
        return JsonResponse({
            'status': 'error',
            'mensaje': 'Este movimiento ya fue autorizado'
        }, status=400)
    
    # Autorizar — usamos QuerySet.update() para bypass de la regla de inmutabilidad
    # (la inmutabilidad en save() protege contra edición de datos de negocio,
    # pero la autorización es un campo de control que debe poder actualizarse)
    MovimientoInventario.objects.filter(pk=movimiento.pk).update(
        autorizado=True,
        autorizado_por=request.user,
        fecha_autorizacion=timezone.now()
    )
    
    return JsonResponse({
        'status': 'success',
        'mensaje': f'Movimiento {movimiento.folio} autorizado exitosamente'
    })


# ==============================================================================
# API: OBTENER LOTES DE UN PRODUCTO
# ==============================================================================
@login_required
def api_lotes_producto(request, producto_id):
    """
    API para obtener lotes disponibles de un producto.
    Devuelve también los datos del producto para el PDV.
    """
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        return JsonResponse({'lotes': [], 'producto': None})

    try:
        producto = Producto.objects.get(id=producto_id, empresa=empresa)
    except Producto.DoesNotExist:
        return JsonResponse({'lotes': [], 'producto': None}, status=404)

    today = date.today()
    lotes = Lote.objects.filter(
        producto=producto,
        cantidad__gt=0
    ).order_by('fecha_caducidad')

    lotes_data = []
    lote_proximo = None  # Primer lote VIGENTE (no vencido) — FEFO
    for l in lotes:
        dias = (l.fecha_caducidad - today).days if l.fecha_caducidad else None
        es_vencido = (l.fecha_caducidad is not None and l.fecha_caducidad < today)
        costo = float(l.costo_adquisicion) if l.costo_adquisicion else 0
        lotes_data.append({
            'id': l.id,
            'numero_lote': l.numero_lote,
            'fecha_caducidad': l.fecha_caducidad.strftime('%Y-%m-%d') if l.fecha_caducidad else None,
            'cantidad': float(l.cantidad),
            'costo_adquisicion': costo,
            'dias_restantes': dias,
            'es_vencido': es_vencido,
        })
        # Solo usar lotes VIGENTES como próximo a despachar (FEFO)
        if lote_proximo is None and not es_vencido:
            lote_proximo = l

    # Stock VIGENTE (excluye vencidos — no se pueden vender)
    stock_vigente = sum(float(l['cantidad']) for l in lotes_data if not l.get('es_vencido'))
    stock_total = sum(float(l['cantidad']) for l in lotes_data)
    lotes_vencidos_count = sum(1 for l in lotes_data if l.get('es_vencido'))

    precio_publico = float(producto.precio_publico)
    costo_lote_proximo = float(lote_proximo.costo_adquisicion) if (lote_proximo and lote_proximo.costo_adquisicion) else 0
    # Alerta de venta a pérdida: precio de venta < costo del lote próximo
    alerta_precio_bajo = (costo_lote_proximo > 0 and precio_publico < costo_lote_proximo)

    producto_data = {
        'id': producto.id,
        'nombre_comercial': producto.nombre,
        'sustancia_activa': producto.sustancia_activa or '',
        'codigo_barras': producto.codigo_barras or '',
        'precio_base': precio_publico,
        'precio_venta': precio_publico,
        'precio_compra': float(producto.precio_compra) if producto.precio_compra else 0,
        'iva_pct': float(producto.iva_porcentaje) if producto.iva_porcentaje else 0,
        'es_antibiotico': producto.es_antibiotico if hasattr(producto, 'es_antibiotico') else False,
        'es_controlado': producto.es_antibiotico if hasattr(producto, 'es_antibiotico') else False,
        'stock': int(producto.stock) if producto.stock else 0,
        'stock_total': stock_vigente,          # Solo stock no vencido
        'stock_total_fisico': stock_total,     # Físico incluyendo vencidos
        'lotes_vencidos_count': lotes_vencidos_count,
        'alerta_precio_bajo': alerta_precio_bajo,
        'costo_lote': costo_lote_proximo,
        'lote_id': lote_proximo.id if lote_proximo else None,
        'numero_lote_proximo': lote_proximo.numero_lote if lote_proximo else '',
        'proxima_caducidad': lote_proximo.fecha_caducidad.strftime('%Y-%m-%d') if lote_proximo and lote_proximo.fecha_caducidad else None,
        # Sin lote vigente = todo está vencido
        'sin_stock_vigente': lote_proximo is None,
    }

    return JsonResponse({'lotes': lotes_data, 'producto': producto_data})


# ==============================================================================
# REPORTE DE VALORIZACIÓN DE INVENTARIO
# ==============================================================================
@login_required
def reporte_valorizacion_inventario(request):
    """
    Reporte financiero de valorización del inventario.
    Calcula el valor total del stock usando costo promedio ponderado.
    """
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada.')
        return redirect('dashboard')
    
    productos = Producto.objects.filter(
        empresa=empresa,
        stock__gt=0
    ).annotate(
        valor_inventario=F('stock') * F('precio_compra')
    ).order_by('-valor_inventario')
    
    # Calcular totales
    total_unidades = productos.aggregate(
        total=Coalesce(Sum('stock'), Value(0), output_field=DecimalField())
    )['total']
    
    total_valor = productos.aggregate(
        total=Coalesce(Sum('valor_inventario'), Value(0), output_field=DecimalField())
    )['total']
    
    # Exportación Excel
    if request.GET.get('formato') == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Valorización'
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Reporte de Valorización — {empresa.nombre}'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f'Fecha: {timezone.now().strftime("%d/%m/%Y %H:%M")}'

        ws.append([])
        headers = ['Producto', 'SKU', 'Categoría', 'Stock', 'Costo Unitario', 'Valor Total']
        ws.append(headers)
        fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        for col in range(1, 7):
            cell = ws.cell(ws.max_row, col)
            cell.fill = fill
            cell.font = Font(color='FFFFFF', bold=True)

        for p in productos:
            # categoria es CharField en Producto (no FK), se accede directamente
            cat_val = p.categoria if isinstance(p.categoria, str) else (
                p.categoria.nombre if hasattr(p.categoria, 'nombre') else str(p.categoria or '')
            )
            ws.append([
                p.nombre,
                p.codigo_barras or '',
                cat_val,
                float(p.stock),
                float(p.precio_compra or 0),
                float(p.valor_inventario or 0),
            ])

        ws.append([])
        ws.append(['', '', '', float(total_unidades), '', float(total_valor)])
        ws.cell(ws.max_row, 4).font = Font(bold=True)
        ws.cell(ws.max_row, 6).font = Font(bold=True)

        for col, width in [('A', 35), ('B', 15), ('C', 20), ('D', 10), ('E', 18), ('F', 18)]:
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="valorizacion_inventario.xlsx"'
        return resp

    return render(request, 'farmacia/reporte_valorizacion.html', {
        'productos': productos,
        'total_unidades': total_unidades,
        'total_valor': total_valor,
        'fecha_reporte': timezone.now()
    })


# ==============================================================================
# REGISTRAR COMPRA CON CÁLCULO DE COSTO PROMEDIO PONDERADO (CPP)
# ==============================================================================
@login_required
@permission_required('farmacia.add_movimientoinventario', raise_exception=True)
def registrar_compra(request):
    """
    Vista para registrar compras a proveedores.
    
    LÓGICA MATEMÁTICA CRÍTICA:
    Al guardar cada producto de la compra:
    1. Crea MovimientoInventario tipo ENTRADA_COMPRA
    2. Crea o actualiza el Lote
    3. RECALCULA EL COSTO PROMEDIO PONDERADO:
       CPP = ((Stock_Anterior * Costo_Anterior) + (Cantidad_Nueva * Costo_Nuevo)) / Stock_Total
    4. Actualiza producto.precio_compra con el nuevo CPP
    5. Actualiza el stock
    
    Esto es LO QUE DEFINE LA UTILIDAD REAL.
    """
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada. Contacte al administrador.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Verificar si es guardar completo o agregar item
        accion = request.POST.get('accion', 'guardar')
        
        if accion == 'guardar_completo':
            try:
                with transaction.atomic():
                    # Obtener datos de la compra
                    proveedor_id = request.POST.get('proveedor')
                    documento_compra = request.POST.get('documento_compra')
                    fecha_compra_str = request.POST.get('fecha_compra')
                    observaciones = request.POST.get('observaciones', '')
                    
                    # Validar proveedor
                    proveedor = get_object_or_404(Proveedor, id=proveedor_id, empresa=empresa)
                    
                    # Obtener items de la sesión
                    items_compra = request.session.get('items_compra_temp', [])
                    
                    if not items_compra:
                        messages.error(request, '⚠️ No hay productos en la compra. Agrega al menos uno.')
                        return redirect('farmacia:registrar_compra')
                    
                    # Procesar cada item
                    total_productos = 0
                    total_valor = Decimal('0.00')
                    
                    for item in items_compra:
                        producto = get_object_or_404(Producto, id=item['producto_id'], empresa=empresa)
                        cantidad = Decimal(str(item['cantidad']))
                        costo_unitario = Decimal(str(item['costo_unitario']))
                        numero_lote = item['numero_lote']
                        fecha_caducidad = datetime.strptime(item['fecha_caducidad'], '%Y-%m-%d').date()
                        marca = item.get('marca', '')
                        
                        # Actualizar marca del producto si se proporciona
                        if marca and marca != producto.marca_laboratorio:
                            producto.marca_laboratorio = marca
                            producto.save(update_fields=['marca_laboratorio'])
                        
                        # ==============================================================
                        # 1. CREAR O ACTUALIZAR LOTE
                        # ==============================================================
                        lote, lote_creado = Lote.objects.get_or_create(
                            producto=producto,
                            numero_lote=numero_lote,
                            defaults={
                                'fecha_caducidad': fecha_caducidad,
                                'cantidad': Decimal('0'),
                                'costo_adquisicion': costo_unitario
                            }
                        )
                        
                        if not lote_creado:
                            # Si el lote ya existía, actualizar su costo (promedio simple)
                            cantidad_anterior = lote.cantidad
                            if cantidad_anterior > 0:
                                lote.costo_adquisicion = (
                                    (cantidad_anterior * lote.costo_adquisicion + cantidad * costo_unitario) /
                                    (cantidad_anterior + cantidad)
                                )
                        
                        # ==============================================================
                        # 2. CREAR MOVIMIENTO EN KARDEX (ESTO ACTUALIZA TODO)
                        # ==============================================================
                        movimiento = MovimientoInventario(
                            empresa=empresa,
                            sucursal=getattr(request.user, 'sucursal', None),
                            producto=producto,
                            lote=lote,
                            tipo_movimiento='ENTRADA_COMPRA',
                            cantidad=cantidad,
                            costo_unitario=costo_unitario,
                            usuario_responsable=request.user,
                            proveedor=proveedor,
                            observaciones=f"Compra a {proveedor.razon_social}. Doc: {documento_compra}",
                            documento_referencia=documento_compra
                        )
                        
                        # ⚠️ IMPORTANTE: El método save() de MovimientoInventario:
                        # - Actualiza lote.cantidad
                        # - Actualiza producto.stock
                        # - RECALCULA producto.precio_compra (CPP)
                        movimiento.save()
                        
                        total_productos += 1
                        total_valor += cantidad * costo_unitario
                    
                    # Limpiar sesión
                    request.session['items_compra_temp'] = []
                    
                    # Registrar en AuditLog
                    from core.models import AuditLog
                    AuditLog.objects.create(
                        empresa=empresa,
                        usuario=request.user,
                        accion=AuditLog.ACCION_CREATE,
                        modelo_afectado='Compra',
                        objeto_id='0',
                        datos_anteriores=None,
                        datos_nuevos={
                            'proveedor': proveedor.razon_social,
                            'documento': documento_compra,
                            'total_productos': total_productos,
                            'total_valor': str(total_valor),
                            'fecha_compra': fecha_compra_str
                        },
                        sucursal=getattr(request.user, 'sucursal', None),
                        ip_address=request.META.get('REMOTE_ADDR'),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
                    )
                    
                    messages.success(
                        request,
                        f'✅ Compra registrada exitosamente. '
                        f'{total_productos} productos agregados al inventario. '
                        f'Valor total: ${total_valor:,.2f}'
                    )
                    
                    return redirect('farmacia:kardex_list')
                    
            except Exception as e:
                messages.error(request, f'❌ Error al registrar compra: {str(e)}')
                return redirect('farmacia:registrar_compra')
    
    # GET: Mostrar formulario
    form_compra = RegistrarCompraForm(empresa=empresa)
    form_detalle = DetalleCompraForm(empresa=empresa)
    
    # Obtener items temporales de la sesión
    items_temp = request.session.get('items_compra_temp', [])
    total_temp = sum(
        Decimal(str(item['cantidad'])) * Decimal(str(item['costo_unitario']))
        for item in items_temp
    )
    
    return render(request, 'farmacia/registrar_compra.html', {
        'form_compra': form_compra,
        'form_detalle': form_detalle,
        'items_temp': items_temp,
        'total_temp': total_temp
    })


# ==============================================================================
# API: AGREGAR PRODUCTO A COMPRA TEMPORAL
# ==============================================================================
@login_required
def api_agregar_producto_compra(request):
    """
    API para agregar productos a la compra temporal (en sesion).
    Soporta multi-lote: se pueden agregar multiples lotes del mismo producto.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            producto_id = data.get('producto_id')
            cantidad = Decimal(str(data.get('cantidad', 0)))
            costo_unitario = Decimal(str(data.get('costo_unitario', 0)))
            numero_lote = data.get('numero_lote', '').strip().upper()
            fecha_caducidad = data.get('fecha_caducidad')
            marca = data.get('marca', '').strip()
            
            # Validaciones
            if not all([producto_id, cantidad > 0, costo_unitario > 0, numero_lote, fecha_caducidad]):
                return JsonResponse({
                    'status': 'error',
                    'mensaje': 'Todos los campos son obligatorios'
                }, status=400)
            
            # Verificar que el producto existe
            producto = get_object_or_404(Producto, id=producto_id, empresa=getattr(request.user, 'empresa', None))
            
            # Agregar a sesion
            if 'items_compra_temp' not in request.session:
                request.session['items_compra_temp'] = []
            
            item = {
                'producto_id': producto.id,
                'producto_nombre': producto.nombre,
                'cantidad': str(cantidad),
                'costo_unitario': str(costo_unitario),
                'subtotal': str(cantidad * costo_unitario),
                'numero_lote': numero_lote,
                'fecha_caducidad': fecha_caducidad,
                'marca': marca or producto.marca_laboratorio or 'GENERICO'
            }
            
            request.session['items_compra_temp'].append(item)
            request.session.modified = True
            
            return JsonResponse({
                'status': 'success',
                'mensaje': f'Lote {numero_lote} de {producto.nombre} agregado',
                'item': item,
                'total_items': len(request.session['items_compra_temp'])
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error: {str(e)}'
            }, status=500)
    
    return JsonResponse({'status': 'error', 'mensaje': 'Metodo no permitido'}, status=405)


# ==============================================================================
# API: AGREGAR MULTIPLES LOTES DE UNA VEZ (MULTI-LOTE)
# ==============================================================================
@login_required
def api_agregar_multi_lote(request):
    """
    API para agregar multiples lotes de un mismo producto a la compra temporal.
    Recibe un array de lotes para un unico producto.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            producto_id = data.get('producto_id')
            lotes = data.get('lotes', [])
            
            if not producto_id or not lotes:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': 'Se requiere producto_id y al menos un lote'
                }, status=400)
            
            producto = get_object_or_404(Producto, id=producto_id, empresa=getattr(request.user, 'empresa', None))
            
            if 'items_compra_temp' not in request.session:
                request.session['items_compra_temp'] = []
            
            items_agregados = []
            for lote_data in lotes:
                cantidad = Decimal(str(lote_data.get('cantidad', 0)))
                costo_unitario = Decimal(str(lote_data.get('costo_unitario', 0)))
                numero_lote = lote_data.get('numero_lote', '').strip().upper()
                fecha_caducidad = lote_data.get('fecha_caducidad')
                marca = lote_data.get('marca', '').strip()
                
                if not all([cantidad > 0, costo_unitario > 0, numero_lote, fecha_caducidad]):
                    continue  # Saltar lotes incompletos
                
                item = {
                    'producto_id': producto.id,
                    'producto_nombre': producto.nombre,
                    'cantidad': str(cantidad),
                    'costo_unitario': str(costo_unitario),
                    'subtotal': str(cantidad * costo_unitario),
                    'numero_lote': numero_lote,
                    'fecha_caducidad': fecha_caducidad,
                    'marca': marca or producto.marca_laboratorio or 'GENERICO'
                }
                
                request.session['items_compra_temp'].append(item)
                items_agregados.append(item)
            
            request.session.modified = True
            
            return JsonResponse({
                'status': 'success',
                'mensaje': f'{len(items_agregados)} lote(s) de {producto.nombre} agregados',
                'items': items_agregados,
                'total_items': len(request.session['items_compra_temp'])
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error: {str(e)}'
            }, status=500)
    
    return JsonResponse({'status': 'error', 'mensaje': 'Metodo no permitido'}, status=405)


# ==============================================================================
# API: ELIMINAR PRODUCTO DE COMPRA TEMPORAL
# ==============================================================================
@login_required
def api_eliminar_producto_compra(request, index):
    """
    API para eliminar un producto de la compra temporal.
    """
    try:
        items = request.session.get('items_compra_temp', [])
        if 0 <= index < len(items):
            item_eliminado = items.pop(index)
            request.session['items_compra_temp'] = items
            request.session.modified = True
            
            return JsonResponse({
                'status': 'success',
                'mensaje': f'Producto {item_eliminado["producto_nombre"]} eliminado'
            })
        else:
            return JsonResponse({
                'status': 'error',
                'mensaje': 'Índice inválido'
            }, status=400)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'mensaje': f'Error: {str(e)}'
        }, status=500)


# ==============================================================================
# CORTE DE CAJA FARMACIA (ARQUEO CIEGO)
# ==============================================================================
@login_required
def corte_caja_farmacia(request):
    """
    Vista para realizar el corte de caja al final del turno.
    
    ARQUEO CIEGO: El cajero NO ve cuánto espera el sistema.
    Solo ingresa el dinero real que tiene.
    
    Al enviar:
    1. Sistema compara Total Declarado vs Total Sistema
    2. Calcula Diferencia (sobrante/faltante)
    3. Genera ticket/PDF inmutable
    4. Registra en AuditLog
    """
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada.')
        return redirect('dashboard')
    usuario = request.user
    
    if request.method == 'POST':
        form = CorteCajaFarmaciaForm(request.POST)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Obtener montos declarados
                    efectivo_declarado = form.cleaned_data['efectivo_declarado']
                    tarjeta_declarada = form.cleaned_data.get('tarjeta_declarada', Decimal('0'))
                    transferencia_declarada = form.cleaned_data.get('transferencia_declarada', Decimal('0'))
                    observaciones = form.cleaned_data.get('observaciones_corte', '')
                    
                    total_declarado = efectivo_declarado + tarjeta_declarada + transferencia_declarada
                    
                    # ============================================================
                    # CALCULAR TOTAL SISTEMA (VENTAS DEL TURNO)
                    # ============================================================
                    # Definir rango del turno (hoy desde las 00:00 hasta ahora)
                    hoy_inicio = datetime.combine(date.today(), time.min)
                    ahora = timezone.now()
                    
                    ventas_turno = Venta.objects.filter(
                        empresa=empresa,
                        fecha__gte=hoy_inicio,
                        fecha__lte=ahora,
                        usuario=usuario  # Solo ventas del cajero actual
                    ).exclude(estado='CANCELADA')
                    
                    # Total esperado por el sistema
                    total_sistema = ventas_turno.aggregate(
                        total=Coalesce(Sum('total'), Value(Decimal('0')), output_field=DecimalField())
                    )['total']
                    
                    # Desglose por método de pago (usando campos multimodales)
                    pagos_desglose = Pago.objects.filter(
                        venta__in=ventas_turno
                    ).aggregate(
                        total_efectivo=Coalesce(Sum('monto_efectivo'), Value(Decimal('0')), output_field=DecimalField()),
                        total_tarjeta=Coalesce(Sum('monto_tarjeta'), Value(Decimal('0')), output_field=DecimalField()),
                        total_transferencia=Coalesce(Sum('monto_transferencia'), Value(Decimal('0')), output_field=DecimalField()),
                    )
                    pagos_efectivo = pagos_desglose['total_efectivo']
                    pagos_tarjeta = pagos_desglose['total_tarjeta']
                    pagos_transferencia = pagos_desglose['total_transferencia']
                    
                    # ============================================================
                    # CALCULAR DIFERENCIAS
                    # ============================================================
                    diferencia_efectivo = efectivo_declarado - pagos_efectivo
                    diferencia_tarjeta = tarjeta_declarada - pagos_tarjeta
                    diferencia_transferencia = transferencia_declarada - pagos_transferencia
                    diferencia_total = total_declarado - total_sistema
                    
                    # Determinar estado
                    if abs(diferencia_total) <= Decimal('1.00'):  # Tolerancia de $1
                        estado = 'CUADRADO'
                        nivel_alerta = 'success'
                    elif diferencia_total > 0:
                        estado = 'SOBRANTE'
                        nivel_alerta = 'warning'
                    else:
                        estado = 'FALTANTE'
                        nivel_alerta = 'danger'
                    
                    # ============================================================
                    # REGISTRAR CORTE EN AUDITLOG (INMUTABLE)
                    # ============================================================
                    from core.models import AuditLog
                    corte_log = AuditLog.objects.create(
                        empresa=empresa,
                        usuario=usuario,
                        accion=AuditLog.ACCION_CREATE,
                        modelo_afectado='CorteCajaFarmacia',
                        objeto_id='0',
                        datos_anteriores=None,
                        datos_nuevos={
                            'fecha_corte': ahora.isoformat(),
                            'turno_inicio': hoy_inicio.isoformat(),
                            'turno_fin': ahora.isoformat(),
                            'num_ventas': ventas_turno.count(),
                            'sistema_total': str(total_sistema),
                            'sistema_efectivo': str(pagos_efectivo),
                            'sistema_tarjeta': str(pagos_tarjeta),
                            'sistema_transferencia': str(pagos_transferencia),
                            'declarado_total': str(total_declarado),
                            'declarado_efectivo': str(efectivo_declarado),
                            'declarado_tarjeta': str(tarjeta_declarada),
                            'declarado_transferencia': str(transferencia_declarada),
                            'diferencia_total': str(diferencia_total),
                            'diferencia_efectivo': str(diferencia_efectivo),
                            'diferencia_tarjeta': str(diferencia_tarjeta),
                            'estado': estado,
                            'observaciones': observaciones
                        },
                        sucursal=getattr(usuario, 'sucursal', None),
                        ip_address=request.META.get('REMOTE_ADDR'),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
                    )
                    
                    # Mostrar resultado en una página de resumen
                    return render(request, 'farmacia/corte_caja_resultado.html', {
                        'corte_id': corte_log.id,
                        'fecha_corte': ahora,
                        'total_ventas': ventas_turno.count(),
                        'sistema_total': total_sistema,
                        'sistema_efectivo': pagos_efectivo,
                        'sistema_tarjeta': pagos_tarjeta,
                        'declarado_total': total_declarado,
                        'declarado_efectivo': efectivo_declarado,
                        'declarado_tarjeta': tarjeta_declarada,
                        'declarado_transferencia': transferencia_declarada,
                        'diferencia_total': diferencia_total,
                        'diferencia_efectivo': diferencia_efectivo,
                        'diferencia_tarjeta': diferencia_tarjeta,
                        'estado': estado,
                        'nivel_alerta': nivel_alerta,
                        'observaciones': observaciones
                    })
                    
            except Exception as e:
                messages.error(request, f'❌ Error al procesar corte de caja: {str(e)}')
    else:
        form = CorteCajaFarmaciaForm()
    
    # Calcular ventas del turno (para mostrar contador, SIN mostrar el total esperado)
    hoy_inicio = datetime.combine(date.today(), time.min)
    ahora = timezone.now()
    
    ventas_turno_count = Venta.objects.filter(
        empresa=empresa,
        fecha__gte=hoy_inicio,
        fecha__lte=ahora,
        usuario=usuario,
    ).exclude(estado='CANCELADA').count()
    
    return render(request, 'farmacia/corte_caja_form.html', {
        'form': form,
        'ventas_turno_count': ventas_turno_count,
        'turno_inicio': hoy_inicio
    })


# ==============================================================================
# GENERAR ETIQUETAS CON CÓDIGO DE BARRAS
# ==============================================================================
@login_required
def generar_etiquetas(request):
    """
    Vista para generar etiquetas con código de barras (Code128).
    
    Flujo:
    1. Seleccionar productos
    2. Configurar formato de etiqueta
    3. Generar PDF con códigos de barras
    4. Descargar para imprimir en impresora Zebra/Dymo
    """
    empresa = getattr(request.user, 'empresa', None)
    if not empresa:
        messages.error(request, 'Usuario no tiene empresa asignada.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = GenerarEtiquetasForm(empresa=empresa, data=request.POST)
        
        if form.is_valid():
            try:
                productos = form.cleaned_data['productos']
                incluir_precio = form.cleaned_data['incluir_precio']
                incluir_caducidad = form.cleaned_data['incluir_caducidad']
                tamaño_etiqueta = form.cleaned_data['tamaño_etiqueta']
                cantidad_por_producto = form.cleaned_data['cantidad_por_producto']
                
                # ============================================================
                # GENERAR PDF CON CÓDIGOS DE BARRAS
                # ============================================================
                from reportlab.pdfgen import canvas
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.units import mm
                from reportlab.graphics.barcode import code128
                from io import BytesIO
                
                buffer = BytesIO()
                
                # Configurar tamaño de página según tipo de etiqueta
                if tamaño_etiqueta == 'zebra_4x6':
                    page_width = 4 * 25.4 * mm  # 4 pulgadas
                    page_height = 6 * 25.4 * mm  # 6 pulgadas
                elif tamaño_etiqueta == 'dymo_2x1':
                    page_width = 2 * 25.4 * mm
                    page_height = 1 * 25.4 * mm
                else:  # A4
                    page_width, page_height = A4
                
                p = canvas.Canvas(buffer, pagesize=(page_width, page_height))
                
                for producto in productos:
                    for i in range(cantidad_por_producto):
                        # Dibujar etiqueta
                        y_position = page_height - 20*mm
                        
                        # Código de barras (usando SKU o ID)
                        codigo = producto.codigo_barras or f"PROD-{producto.id:06d}"
                        barcode = code128.Code128(codigo, barHeight=15*mm, barWidth=0.8)
                        barcode.drawOn(p, 10*mm, y_position - 15*mm)
                        
                        # Nombre del producto
                        p.setFont("Helvetica-Bold", 12)
                        p.drawString(10*mm, y_position - 20*mm, producto.nombre[:40])
                        
                        # Precio (si se solicita)
                        if incluir_precio and producto.precio_publico:
                            p.setFont("Helvetica", 18)
                            p.drawString(10*mm, y_position - 28*mm, f"${producto.precio_publico:,.2f}")
                        
                        # Caducidad (si se solicita)
                        if incluir_caducidad:
                            lote_proximo = producto.lotes.filter(cantidad__gt=0).order_by('fecha_caducidad').first()
                            if lote_proximo:
                                p.setFont("Helvetica", 8)
                                p.drawString(10*mm, y_position - 35*mm, f"Cad: {lote_proximo.fecha_caducidad.strftime('%m/%Y')}")
                        
                        # Nueva página para siguiente etiqueta
                        p.showPage()
                
                p.save()
                buffer.seek(0)
                
                # Retornar PDF
                response = HttpResponse(buffer, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="etiquetas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
                
                return response
                
            except Exception as e:
                messages.error(request, f'❌ Error al generar etiquetas: {str(e)}')
    else:
        form = GenerarEtiquetasForm(empresa=empresa)
    
    return render(request, 'farmacia/generar_etiquetas.html', {
        'form': form
    })


# ==============================================================================
# IMPORTAR VISTAS DE SOPORTE OPERATIVO V5.0
# ==============================================================================
from farmacia.views.soporte import (
    # Devoluciones
    buscar_venta_para_devolucion,
    procesar_devolucion,
    dashboard_devoluciones,
    autorizar_devolucion,
    
    # Apertura de Caja
    verificar_apertura_caja,
    abrir_caja,
    
    # Antibióticos (COFEPRIS)
    validar_venta_antibiotico,
    reporte_cofepris,
    
    # Entrada Express
    entrada_express,
)

# Importar vistas de Semáforo de Caducidad (ya existente)
from farmacia.views.semaforo import dashboard_semaforo_caducidad, dashboard_stock_critico

