"""
IMPORTACION DE INVENTARIO DESDE EXCEL - PRISLAB V5
====================================================
Lee el archivo Excel exportado del sistema POS y carga productos + lotes
en la base de datos SIN eliminar datos existentes (update_or_create).

Columnas del Excel que se procesan:
  Nombre del Producto, Identificador (No Cambiar), Es un Servicio, Se Vende,
  Descripcion, Categoria, Marca, Unidad de Venta, Codigo de Barras,
  Usa Lotes, Lote, Fabricacion del Lote, Caducidad del Lote,
  Utiliza Stock, Stock Total, Stock Minimo, Precio Publico,
  Precio PERSONAL, Costo, IVA, Receta Medica

Reglas especiales:
  - Productos con "Usa Lotes = No": se crea un lote sintetico para que
    el algoritmo PEPS del PDV pueda encontrar y vender el producto.
  - Productos sin fecha de caducidad: se asigna 2099-12-31 (sin vencimiento).
  - Lotes con fecha ya vencida (6 en el Excel): se importan con cantidad=0.
  - Productos con codigo_barras vacio: se genera uno desde el identificador.
  - Se agrupan filas por Identificador para manejar multiples lotes por producto.
  - NO se borra ningun dato existente.

Uso:
  python manage.py importar_excel_inventario
  python manage.py importar_excel_inventario /ruta/al/archivo.xlsx
  python manage.py importar_excel_inventario --dry-run   (solo reporta, no guarda)
"""
import os
import openpyxl
import datetime
from decimal import Decimal, InvalidOperation
from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction
from django.db.models import Model as DjangoModel, Value
from django.utils import timezone

from core.models import Producto, Lote, Empresa, Sucursal


# ---------------------------------------------------------------------------
# Utilidades de conversion
# ---------------------------------------------------------------------------
FECHA_SIN_VENCIMIENTO = datetime.date(2099, 12, 31)


def _dec(valor, default=Decimal("0.00")):
    if valor is None:
        return default
    try:
        s = str(valor).replace(",", "").replace("$", "").strip()
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return default


def _int(valor, default=0):
    if valor is None:
        return default
    try:
        return int(float(str(valor).replace(",", "").strip()))
    except (ValueError, TypeError):
        return default


def _bool(valor):
    if valor is None:
        return False
    return str(valor).strip().lower() in ("si", "sí", "yes", "1", "true")


def _fecha(valor):
    """Convierte datetime/date/str a date. Devuelve None si no se puede."""
    if valor is None:
        return None
    if hasattr(valor, "date"):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    try:
        return datetime.datetime.strptime(str(valor)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _categoria(raw):
    """Mapea la categoria del Excel al choice del modelo."""
    raw = str(raw or "").strip().upper()
    mapa = {
        "ANTIBIOTICO":  "ANTIBIOTICO",
        "ANTIBIÓTICO":  "ANTIBIOTICO",
        "PATENTE":      "PATENTE",
        "GENERICO":     "GENERICO",
        "GENÉRICO":     "GENERICO",
        "CURACION":     "CURACION",
        "CURACIÓN":     "CURACION",
        "MEDICAMENTO":  "PATENTE",
        "OTRO":         "OTRO",
        "OTROS":        "OTRO",
        "SERVICIO":     "OTRO",
        "SUPLEMENTO":   "OTRO",
    }
    return mapa.get(raw, "OTRO")


def _numero_lote_default(identificador):
    slug = str(identificador or "").replace("-", "")[:12].upper()
    return f"IMP-{slug or 'STOCK'}"


def _barcode_desde_identificador(identificador, empresa_id):
    """Genera un codigo de barras unico cuando el Excel no trae uno."""
    slug = str(identificador or "").replace("-", "")[:16].upper()
    return f"PRIS-{empresa_id}-{slug}"


# ---------------------------------------------------------------------------
# Guardar Lote SALTANDO full_clean() para importacion historica
# (evita que lotes vencidos del Excel bloqueen la importacion)
# ---------------------------------------------------------------------------
def _guardar_lote_sin_validar(lote: Lote):
    """
    Guarda el objeto Lote usando el save() del padre (Django base),
    evitando la validacion de caducidad que bloquea lotes vencidos.
    A continuacion actualiza el stock total del Producto manualmente.
    """
    DjangoModel.save(lote)
    # Actualizar stock del producto sumando la cantidad de este lote
    if lote.cantidad and lote.cantidad > 0:
        from django.db.models import F
        Producto.objects.filter(pk=lote.producto_id).update(
            stock=F("stock") + lote.cantidad
        )


# ---------------------------------------------------------------------------
# COMANDO PRINCIPAL
# ---------------------------------------------------------------------------
class Command(BaseCommand):
    help = "Importa productos y lotes desde el Excel de inventario PRISLAB"

    def add_arguments(self, parser):
        parser.add_argument(
            "archivo",
            nargs="?",
            default=None,
            help="Ruta al archivo .xlsx (default: auto-detecta en Downloads o raiz)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Solo reporta lo que haria, sin guardar nada",
        )
        parser.add_argument(
            "--empresa-id",
            type=int,
            default=None,
            help="ID de la empresa (tenant). Obligatorio salvo --empresa (nombre o ID como texto).",
        )
        parser.add_argument(
            "--empresa",
            default=None,
            help="Nombre o ID numérico como texto (alternativa a --empresa-id).",
        )
        parser.add_argument(
            "--reset-stock",
            action="store_true",
            help="Pone el stock del producto a 0 antes de importar sus lotes (re-calcula)",
        )

    def handle(self, *args, **options):
        archivo   = options["archivo"]
        dry_run   = options["dry_run"]
        reset_stk = options["reset_stock"]

        # ---- Localizar archivo ----
        if not archivo:
            candidatos = [
                Path(r"C:\Users\jonil\Downloads\Productos-2026-02-22-16-01.xlsx"),
                Path("Productos-2026-02-22-16-01.xlsx"),
                Path("inventario.xlsx"),
            ]
            for c in candidatos:
                if c.exists():
                    archivo = str(c)
                    break

        if not archivo or not Path(archivo).exists():
            self.stdout.write(self.style.ERROR(
                f"No se encontro el archivo Excel. Pasa la ruta como argumento."
            ))
            return

        self.stdout.write(self.style.SUCCESS(f"[INICIO] Leyendo: {archivo}"))
        if dry_run:
            self.stdout.write(self.style.WARNING("  [DRY-RUN] No se guardara nada."))

        # ---- Empresa (sin fallback a primera fila de la tabla) ----
        empresa = None
        eid = options.get("empresa_id")
        if eid:
            empresa = Empresa.objects.filter(pk=eid).first()
        if not empresa and options["empresa"]:
            raw = str(options["empresa"]).strip()
            try:
                empresa = Empresa.objects.get(pk=int(raw))
            except (ValueError, Empresa.DoesNotExist):
                empresa = Empresa.objects.filter(nombre__icontains=raw).first()
        if not empresa:
            self.stdout.write(
                self.style.ERROR(
                    "Indique --empresa-id=<pk> o --empresa (nombre o ID). "
                    "No se usa empresa implícita por seguridad multi-tenant."
                )
            )
            return
        self.stdout.write(f"  Empresa: {empresa.nombre} (ID={empresa.id})")

        # ---- Leer Excel ----
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        headers_raw = [str(c) if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        COL = {h.strip(): i for i, h in enumerate(headers_raw) if h.strip()}

        def col(row, nombre, default=None):
            # Buscar con y sin espacios finales (el Excel puede tener "Stock Total ")
            idx = COL.get(str(nombre).strip())
            if idx is None:
                return default
            v = row[idx]
            return v if v is not None else default

        # ---- Agrupar filas por identificador ----
        grupos = defaultdict(list)
        total_filas = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            ident = str(col(row, "Identificador (No Cambiar)", "") or row[0]).strip().lower()
            grupos[ident].append(row)
            total_filas += 1

        self.stdout.write(f"  Filas de datos: {total_filas} | Productos unicos: {len(grupos)}")

        # ---- Contadores ----
        prod_creados = prod_actualizados = lote_creados = lote_ya_existe = 0
        errores = []
        hoy = datetime.date.today()

        # ---- Importar ----
        for ident, filas in grupos.items():
            # Tomar datos del producto de la primera fila del grupo
            fila_prod = filas[0]

            nombre       = str(col(fila_prod, "Nombre del Producto", ident)).strip()
            es_servicio  = _bool(col(fila_prod, "Es un Servicio"))
            usa_lotes    = _bool(col(fila_prod, "Usa Lotes"))
            categoria    = _categoria(col(fila_prod, "Categoría"))
            marca        = str(col(fila_prod, "Marca", "GENÉRICO") or "GENÉRICO").strip()
            unidad       = str(col(fila_prod, "Unidad de Venta", "Unidad") or "Unidad").strip()
            barcode_raw  = col(fila_prod, "Código de Barras")
            descripcion  = str(col(fila_prod, "Descripción", "") or "").strip()
            precio_pub   = _dec(col(fila_prod, "Precio Público", 0))
            costo        = _dec(col(fila_prod, "Costo", 0))
            iva_pct      = _dec(col(fila_prod, "IVA", 0))
            stk_min      = _int(col(fila_prod, "Stock Mínimo ", 0))
            receta       = _bool(col(fila_prod, "Receta Médica"))

            # Codigo de barras: generar si falta
            barcode = str(barcode_raw).strip() if barcode_raw else None
            if not barcode:
                barcode = _barcode_desde_identificador(ident, empresa.id)

            # Resolver conflictos de barcode (si otro producto ya tiene ese barcode)
            barcode_final = barcode
            if Producto.objects.filter(codigo_barras=barcode).exclude(
                empresa=empresa
            ).exists():
                barcode_final = f"{barcode}-{empresa.id}"

            # ---- Crear o actualizar Producto ----
            campos_producto = dict(
                nombre               = nombre,
                marca_laboratorio    = marca[:150],
                forma_farmaceutica   = unidad[:100],
                concentracion        = "N/A",
                presentacion         = "1",
                categoria            = categoria,
                precio_compra        = costo,
                precio_publico       = precio_pub,
                iva_porcentaje       = iva_pct,
                stock_minimo         = stk_min if stk_min >= 0 else 0,
                es_servicio          = es_servicio,
                es_antibiotico       = receta,
                sustancia_activa     = descripcion[:255] if descripcion else None,
            )
            # stock: se restablecera al sumar lotes mas adelante
            if reset_stk:
                campos_producto["stock"] = 0

            try:
                if dry_run:
                    existe = Producto.objects.filter(empresa=empresa, codigo_barras=barcode_final).exists()
                    if existe:
                        prod_actualizados += 1
                    else:
                        prod_creados += 1
                    producto_obj = None
                else:
                    with transaction.atomic():
                        producto_obj, creado = Producto.objects.update_or_create(
                            empresa       = empresa,
                            codigo_barras = barcode_final,
                            defaults      = campos_producto,
                        )
                        if reset_stk and not creado:
                            Producto.objects.filter(pk=producto_obj.pk).update(stock=0)
                    if creado:
                        prod_creados += 1
                    else:
                        prod_actualizados += 1
            except Exception as e:
                errores.append(f"[PROD] {nombre}: {e}")
                continue

            # ---- Crear Lotes ----
            for fila_lote in filas:
                # Leer datos del lote de ESTA fila
                num_lote    = str(col(fila_lote, "Lote", "") or "").strip()
                fab_raw     = col(fila_lote, "Fabricación del Lote")
                cad_raw     = col(fila_lote, "Caducidad del Lote")
                stock_total = _int(col(fila_lote, "Stock Total ", 0))

                # Resolver numero de lote
                if not num_lote:
                    if usa_lotes:
                        num_lote = _numero_lote_default(ident)
                    else:
                        num_lote = f"GEN-{ident[:12].upper().replace('-','')}"

                # Resolver fecha de caducidad
                fecha_cad = _fecha(cad_raw)
                if fecha_cad is None:
                    fecha_cad = FECHA_SIN_VENCIMIENTO  # Sin vencimiento = 2099-12-31

                # Lotes ya vencidos: importar con cantidad 0 para preservar el registro
                cantidad_lote = stock_total
                if fecha_cad < hoy:
                    cantidad_lote = 0

                fecha_fab = _fecha(fab_raw)

                if dry_run:
                    lote_creados += 1
                    continue

                if producto_obj is None:
                    continue

                try:
                    with transaction.atomic():
                        lote_existe = Lote.objects.filter(
                            producto    = producto_obj,
                            numero_lote = num_lote,
                        ).first()

                        if lote_existe:
                            # Actualizar usando UPDATE SQL directo (evita override de save)
                            from django.db.models import F
                            cambios = {}
                            if lote_existe.cantidad != cantidad_lote:
                                cambios["cantidad"] = cantidad_lote
                            if lote_existe.costo_adquisicion != costo and costo > 0:
                                cambios["costo_adquisicion"] = costo
                            if lote_existe.fecha_caducidad != fecha_cad:
                                cambios["fecha_caducidad"] = fecha_cad
                            if cambios:
                                Lote.objects.filter(pk=lote_existe.pk).update(**cambios)

                            # Sincronizar stock del producto
                            if reset_stk:
                                # Stock fue reseteado a 0: sumar cantidad completa
                                Producto.objects.filter(pk=producto_obj.pk).update(
                                    stock=F("stock") + cantidad_lote
                                )
                            elif "cantidad" in cambios:
                                # Sin reset: solo sumar la diferencia
                                diff = cantidad_lote - lote_existe.cantidad
                                Producto.objects.filter(pk=producto_obj.pk).update(
                                    stock=F("stock") + diff
                                )
                            lote_ya_existe += 1
                        else:
                            nuevo_lote = Lote(
                                producto         = producto_obj,
                                numero_lote      = num_lote,
                                fecha_fabricacion= fecha_fab,
                                fecha_caducidad  = fecha_cad,
                                cantidad         = cantidad_lote,
                                costo_adquisicion= costo if costo > 0 else Decimal("0.01"),
                            )
                            _guardar_lote_sin_validar(nuevo_lote)
                            lote_creados += 1

                except Exception as e:
                    errores.append(f"[LOTE] {nombre} / {num_lote}: {e}")

        wb.close()

        # ---- Reporte final ----
        sep = "=" * 60
        self.stdout.write(f"\n{sep}")
        self.stdout.write(self.style.SUCCESS("  IMPORTACION COMPLETADA"))
        self.stdout.write(sep)
        self.stdout.write(f"  Productos creados      : {prod_creados}")
        self.stdout.write(f"  Productos actualizados : {prod_actualizados}")
        self.stdout.write(f"  Lotes creados          : {lote_creados}")
        self.stdout.write(f"  Lotes ya existian      : {lote_ya_existe}")
        self.stdout.write(f"  Errores                : {len(errores)}")
        if errores:
            self.stdout.write(self.style.WARNING("\n  ERRORES:"))
            for e in errores[:20]:
                self.stdout.write(f"    {e}")
        if dry_run:
            self.stdout.write(self.style.WARNING("\n  [DRY-RUN] Nada fue guardado."))
        else:
            # Recalcular Producto.stock sumando SOLO lotes con cantidad > 0
            # (sincroniza el campo con la realidad de los lotes)
            self.stdout.write("\n  Recalculando campo stock desde lotes...")
            from django.db.models import Sum
            from django.db.models.functions import Coalesce
            stk_corregidos = 0
            for p in Producto.objects.filter(empresa=empresa):
                agg = Lote.objects.filter(producto=p, cantidad__gt=0).aggregate(
                    total=Coalesce(Sum("cantidad"), Value(0))
                )
                stock_real = agg["total"]
                if p.stock != stock_real:
                    Producto.objects.filter(pk=p.pk).update(stock=stock_real)
                    stk_corregidos += 1
            self.stdout.write(f"    Productos con stock corregido: {stk_corregidos}")

            # Verificacion final
            total_prods = Producto.objects.filter(empresa=empresa).count()
            total_lotes = Lote.objects.filter(producto__empresa=empresa).count()
            lotes_con_stock = Lote.objects.filter(
                producto__empresa=empresa, cantidad__gt=0
            ).count()
            prods_con_stock = Producto.objects.filter(empresa=empresa, stock__gt=0).count()
            self.stdout.write(f"\n  Estado final en DB:")
            self.stdout.write(f"    Productos en DB        : {total_prods}")
            self.stdout.write(f"    Productos con stock    : {prods_con_stock}")
            self.stdout.write(f"    Lotes en DB            : {total_lotes}")
            self.stdout.write(f"    Lotes con stock activo : {lotes_con_stock}")
        self.stdout.write(sep)
