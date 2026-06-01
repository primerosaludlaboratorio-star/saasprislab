"""
COMANDO: Carga Masiva de Inventario de Farmacia desde Excel
Soporta formatos .xlsx y .xls

Uso: python manage.py cargar_inventario_excel <archivo.xlsx>
"""
import os
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.db import transaction
from core.models import Producto, Empresa, Sucursal
import openpyxl


class Command(BaseCommand):
    help = 'Carga inventario de farmacia desde archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta al archivo Excel con el inventario'
        )
        parser.add_argument(
            '--skip-header',
            action='store_true',
            help='Saltar la primera fila (encabezados)'
        )
        parser.add_argument(
            '--empresa-id',
            type=int,
            default=1,
            help='ID de la empresa (default: 1)'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        skip_header = options['skip_header']
        empresa_id = options['empresa_id']

        # Validar archivo
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'[ERROR] Archivo no encontrado: {archivo}'))
            return

        # Obtener empresa
        try:
            empresa = Empresa.objects.get(id=empresa_id)
        except Empresa.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'[ERROR] Empresa con ID {empresa_id} no existe'))
            return

        sucursal = Sucursal.objects.filter(empresa=empresa).first()

        self.stdout.write('=' * 80)
        self.stdout.write(self.style.SUCCESS('CARGA MASIVA DE INVENTARIO - FARMACIA'))
        self.stdout.write('=' * 80)
        self.stdout.write(f'Archivo: {archivo}')
        self.stdout.write(f'Empresa: {empresa.nombre}')
        self.stdout.write(f'Sucursal: {sucursal.nombre if sucursal else "Sin sucursal"}')
        self.stdout.write('')

        # Cargar Excel
        try:
            workbook = openpyxl.load_workbook(archivo, data_only=True)
            sheet = workbook.active
            self.stdout.write(f'[OK] Excel cargado: {sheet.title}')
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'[ERROR] No se pudo abrir el Excel: {e}'))
            return

        # Detectar columnas automáticamente
        self.stdout.write('\n[DETECTANDO COLUMNAS...]')
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # Mapeo flexible de columnas
        columnas = self._detectar_columnas(header_row)
        
        if not columnas:
            self.stdout.write(self.style.WARNING('[ADVERTENCIA] No se detectaron columnas automáticamente'))
            self.stdout.write('Usando configuración por defecto (índices de columna)')
            # Configuración manual por índice
            columnas = {
                'codigo': 0,  # Columna A
                'nombre': 1,  # Columna B
                'laboratorio': 2,  # Columna C
                'precio_compra': 3,  # Columna D
                'precio_venta': 4,  # Columna E
                'stock': 5,  # Columna F
            }
        else:
            self.stdout.write('[OK] Columnas detectadas:')
            for campo, indice in columnas.items():
                self.stdout.write(f'  - {campo}: Columna {indice + 1} ({header_row[indice]})')

        # Procesar filas
        self.stdout.write('\n[PROCESANDO PRODUCTOS...]')
        creados = 0
        actualizados = 0
        errores = []

        start_row = 2 if skip_header else 1

        with transaction.atomic():
            for row_num, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
                try:
                    # Extraer valores según mapeo
                    codigo = self._get_value(row, columnas.get('codigo'))
                    nombre = self._get_value(row, columnas.get('nombre'))
                    
                    if not codigo or not nombre:
                        continue  # Saltar filas vacías

                    # Datos opcionales
                    laboratorio = self._get_value(row, columnas.get('laboratorio'), 'GENERICO')
                    sustancia = self._get_value(row, columnas.get('sustancia'), '')
                    forma = self._get_value(row, columnas.get('forma'), 'Tabletas')
                    concentracion = self._get_value(row, columnas.get('concentracion'), 'N/A')
                    presentacion = self._get_value(row, columnas.get('presentacion'), '1')
                    
                    # Precios
                    precio_compra = self._to_decimal(self._get_value(row, columnas.get('precio_compra'), 0))
                    precio_venta = self._to_decimal(self._get_value(row, columnas.get('precio_venta'), 0))
                    stock = int(self._get_value(row, columnas.get('stock'), 0) or 0)
                    
                    # Clasificación
                    es_antibiotico = 'antibiotic' in str(nombre).lower() or 'antibiotic' in str(sustancia).lower()
                    
                    # Crear o actualizar producto
                    producto, created = Producto.objects.update_or_create(
                        codigo_barras=str(codigo),
                        defaults={
                            'empresa': empresa,
                            'sucursal': sucursal,
                            'nombre': nombre,
                            'sustancia_activa': sustancia,
                            'marca_laboratorio': laboratorio,
                            'forma_farmaceutica': forma,
                            'concentracion': concentracion,
                            'presentacion': presentacion,
                            'precio_compra': precio_compra,
                            'precio_publico': precio_venta,
                            'stock': stock,
                            'iva_porcentaje': Decimal('16.00'),
                            'clasificacion_sanitaria': 'IV' if es_antibiotico else 'VI',
                            'categoria': 'ANTIBIOTICO' if es_antibiotico else 'GENERICO',
                            'es_antibiotico': es_antibiotico,
                            'es_servicio': False,
                        }
                    )

                    if created:
                        creados += 1
                        if creados % 50 == 0:
                            self.stdout.write(f'  [{creados} productos creados...]')
                    else:
                        actualizados += 1

                except Exception as e:
                    error_msg = f'Fila {row_num}: {e}'
                    errores.append(error_msg)
                    if len(errores) <= 10:  # Solo mostrar primeros 10 errores
                        self.stdout.write(self.style.WARNING(f'  [!] {error_msg}'))

        # Reporte final
        self.stdout.write('\n' + '=' * 80)
        self.stdout.write(self.style.SUCCESS('CARGA COMPLETADA'))
        self.stdout.write('=' * 80)
        self.stdout.write(f'[OK] Productos creados: {creados}')
        self.stdout.write(f'[OK] Productos actualizados: {actualizados}')
        
        if errores:
            self.stdout.write(self.style.WARNING(f'[!] Errores encontrados: {len(errores)}'))
            if len(errores) > 10:
                self.stdout.write(f'  (Mostrando solo los primeros 10)')

        self.stdout.write('\n[EXITO] Inventario cargado correctamente')
        self.stdout.write('=' * 80)

    def _detectar_columnas(self, header_row):
        """Detecta automáticamente las columnas del Excel"""
        columnas = {}
        
        for idx, col_name in enumerate(header_row):
            if not col_name:
                continue
            
            col_lower = str(col_name).lower().strip()
            
            # Código/Clave
            if any(k in col_lower for k in ['codigo', 'clave', 'sku', 'barras']):
                columnas['codigo'] = idx
            # Nombre
            elif any(k in col_lower for k in ['nombre', 'producto', 'descripcion']):
                columnas['nombre'] = idx
            # Laboratorio
            elif any(k in col_lower for k in ['laboratorio', 'marca', 'fabricante']):
                columnas['laboratorio'] = idx
            # Sustancia
            elif any(k in col_lower for k in ['sustancia', 'generico', 'activo']):
                columnas['sustancia'] = idx
            # Forma
            elif any(k in col_lower for k in ['forma', 'presentacion', 'tipo']):
                columnas['forma'] = idx
            # Concentración
            elif any(k in col_lower for k in ['concentracion', 'dosis']):
                columnas['concentracion'] = idx
            # Precio Compra
            elif any(k in col_lower for k in ['compra', 'costo', 'adquisicion']):
                columnas['precio_compra'] = idx
            # Precio Venta
            elif any(k in col_lower for k in ['venta', 'publico', 'precio']) and 'compra' not in col_lower:
                columnas['precio_venta'] = idx
            # Stock
            elif any(k in col_lower for k in ['stock', 'existencia', 'inventario', 'cantidad']):
                columnas['stock'] = idx
        
        return columnas if columnas else None

    def _get_value(self, row, index, default=''):
        """Obtiene valor de una fila de Excel de forma segura"""
        if index is None or index >= len(row):
            return default
        value = row[index]
        return value if value is not None else default

    def _to_decimal(self, value):
        """Convierte valor a Decimal de forma segura"""
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace('$', '').strip()
            return Decimal(str(value))
        except:
            return Decimal('0.00')
